import sys
import csv
import os
import shutil
import platform
import ctypes
import requests
from datetime import datetime
from collections import defaultdict
from PIL import Image
try:
    from pyzbar.pyzbar import decode as pyzbar_decode
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False
    def pyzbar_decode(img): return []
import openpyxl
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QSplashScreen, QStyledItemDelegate,
                             QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem,
                             QLineEdit, QHeaderView, QAbstractItemView,
                             QDialog, QFormLayout, QDoubleSpinBox, QMessageBox,
                             QFileDialog, QComboBox, QTabWidget, QLabel, QDateEdit,
                             QListWidget, QListWidgetItem, QSizeGrip, QCheckBox)
from PyQt6.QtCore import Qt, QDate, QTimer, QSettings, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QKeySequence, QIcon, QShortcut

# ─────────────────────────────────────────
#  SUNUCU AYARI — Render URL'nizi buraya girin
# ─────────────────────────────────────────
API_URL = "https://vasak-erp-backend.speedyhck.workers.dev"
YEDEK_DOSYA = "stockflow_yedek.json"

# ─────────────────────────────────────────
#  TEMALAR
# ─────────────────────────────────────────
TEMALAR = {
    "Maviş": {
        "bg": "#181825", "surface": "#1E1E2E", "surface2": "#313244",
        "border": "#45475A", "text": "#CDD6F4", "muted": "#A6ADC8",
        "accent": "#FAB387", "accent2": "#F5E0DC",
        "red": "#F38BA8", "green": "#A6E3A1", "yellow": "#F9E2AF",
        "blue": "#89B4FA", "purple": "#CBA6F7",
        "tab_bg": "#1E1E2E", "tab_sel": "#313244",
        "baslik_bg": "#11111B", "baslik_border": "#313244",
    },
    "Gece": {
        "bg": "#0D1117", "surface": "#161B22", "surface2": "#21262D",
        "border": "#30363D", "text": "#E6EDF3", "muted": "#8B949E",
        "accent": "#58A6FF", "accent2": "#79C0FF",
        "red": "#F85149", "green": "#3FB950", "yellow": "#D29922",
        "blue": "#58A6FF", "purple": "#BC8CFF",
        "tab_bg": "#161B22", "tab_sel": "#21262D",
        "baslik_bg": "#0D1117", "baslik_border": "#30363D",
    },
    "Güneş": {
        "bg": "#FFF8F0", "surface": "#FFFFFF", "surface2": "#F5F5F5",
        "border": "#E0E0E0", "text": "#2D2D2D", "muted": "#757575",
        "accent": "#FF6B35", "accent2": "#FF8C42",
        "red": "#E53935", "green": "#43A047", "yellow": "#F9A825",
        "blue": "#1E88E5", "purple": "#8E24AA",
        "tab_bg": "#F5F5F5", "tab_sel": "#FFFFFF",
        "baslik_bg": "#FF6B35", "baslik_border": "#FF8C42",
    },
    "Orman": {
        "bg": "#1A2416", "surface": "#1E2D1A", "surface2": "#2D4A25",
        "border": "#3D6B30", "text": "#D4EAC8", "muted": "#8FBF7F",
        "accent": "#6BCB4A", "accent2": "#8FE070",
        "red": "#E05A4A", "green": "#6BCB4A", "yellow": "#D4B840",
        "blue": "#4A9FD4", "purple": "#9F7FD4",
        "tab_bg": "#1E2D1A", "tab_sel": "#2D4A25",
        "baslik_bg": "#111A0D", "baslik_border": "#3D6B30",
    },
}
AKTIF_TEMA = "Maviş"
UYGULAMA_VERSIYON = "v3.05"
GITHUB_VERSIYON_URL = "https://raw.githubusercontent.com/Mirza1293/vasak-erp-backend/main/version.txt"
GITHUB_RELEASE_URL = "https://github.com/Mirza1293/vasak-erp-backend/releases/latest/download/StockFlow_v15.exe"

# ─────────────────────────────────────────
#  API İSTEMCİSİ
# ─────────────────────────────────────────
class ApiIstemci:
    def __init__(self):
        self.token = None
        self.oturum = requests.Session()
        self.oturum.headers.update({"Content-Type": "application/json"})

    def giris_yap(self, sifre: str) -> bool:
        try:
            r = self.oturum.post(f"{API_URL}/api/giris", json={"sifre": sifre}, timeout=15)
            if r.status_code == 200:
                self.token = r.json()["token"]
                self.oturum.headers["Authorization"] = f"Bearer {self.token}"
                return True
        except Exception:
            pass
        return False

    def urunleri_getir(self) -> list:
        try:
            r = self.oturum.get(f"{API_URL}/api/urunler", timeout=15)
            if r.status_code == 200:
                return r.json().get("urunler", [])
        except requests.exceptions.ConnectionError:
            raise ConnectionError("Sunucuya bağlanılamadı. İnternet bağlantınızı kontrol edin.")
        return []

    def urun_ekle(self, veri: dict) -> dict:
        r = self.oturum.post(f"{API_URL}/api/urunler", json=veri, timeout=15)
        if r.status_code in (200, 201):
            return r.json()
        raise Exception(r.json().get("detail", "Ürün eklenemedi."))

    def urun_guncelle(self, urun_id: int, guncelleme: dict) -> bool:
        try:
            r = self.oturum.put(f"{API_URL}/api/urunler/{urun_id}", json=guncelleme, timeout=15)
            return r.status_code == 200
        except Exception:
            return False

    def urun_sil(self, urun_id: int) -> bool:
        try:
            r = self.oturum.delete(f"{API_URL}/api/urunler/{urun_id}", timeout=15)
            return r.status_code == 200
        except Exception:
            return False

    def analiz_getir(self) -> dict:
        try:
            r = self.oturum.get(f"{API_URL}/api/analiz", timeout=15)
            if r.status_code == 200:
                return r.json()
        except Exception:
            pass
        return {}

# ─────────────────────────────────────────
#  ARKA PLANDA VERİ YÜKLEME (UI donmasın)
# ─────────────────────────────────────────
class VeriYukleyici(QThread):
    veri_hazir = pyqtSignal(list)
    hata_olustu = pyqtSignal(str)

    def __init__(self, api: ApiIstemci):
        super().__init__()
        self.api = api

    def run(self):
        try:
            veriler = self.api.urunleri_getir()
            self.veri_hazir.emit(veriler)
        except ConnectionError as e:
            self.hata_olustu.emit(str(e))


# ─────────────────────────────────────────
#  GÜNCELLEME KONTROLÜ
# ─────────────────────────────────────────
class GuncellemKontrolcu(QThread):
    guncelleme_var = pyqtSignal(str)
    guncelleme_yok = pyqtSignal()
    hata = pyqtSignal(str)

    def run(self):
        try:
            r = requests.get(GITHUB_VERSIYON_URL, timeout=5)
            if r.status_code == 200:
                uzak_versiyon = r.text.strip()
                if uzak_versiyon != UYGULAMA_VERSIYON:
                    self.guncelleme_var.emit(uzak_versiyon)
                else:
                    self.guncelleme_yok.emit()
        except Exception as e:
            self.hata.emit(str(e))


class GuincellemeIndirici(QThread):
    ilerleme = pyqtSignal(int)
    tamamlandi = pyqtSignal(str)
    hata = pyqtSignal(str)

    def __init__(self, url, hedef):
        super().__init__()
        self.url = url
        self.hedef = hedef

    def run(self):
        try:
            r = requests.get(self.url, stream=True, timeout=60)
            toplam = int(r.headers.get('content-length', 0))
            indirilen = 0
            with open(self.hedef, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        indirilen += len(chunk)
                        if toplam > 0:
                            self.ilerleme.emit(int(indirilen * 100 / toplam))
            self.tamamlandi.emit(self.hedef)
        except Exception as e:
            self.hata.emit(str(e))


# ─────────────────────────────────────────
#  YARDIMCI FONKSIYON
# ─────────────────────────────────────────
def kaynak_yolunu_bul(goreceli_yol):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, goreceli_yol)

def uygulama_ikonu_yukle():
    """Vaşak ikonunu yükle — dosya varsa dosyadan, yoksa gömülü base64'ten"""
    from PyQt6.QtGui import QIcon, QPixmap
    from PyQt6.QtCore import QByteArray
    import base64
    # Önce dosyadan dene
    ico_yol = kaynak_yolunu_bul("vasak_icon.ico")
    if os.path.exists(ico_yol):
        return QIcon(ico_yol)
    png_yol = kaynak_yolunu_bul("vasak_128.png")
    if os.path.exists(png_yol):
        return QIcon(png_yol)
    # Gömülü ikon (32px base64)
    try:
        ICON_B64 = "iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAAHFklEQVR4nM1XbXBU1Rl+3nPvfifZfJEYkoWSgCSLNKWAhRTYSMJASkaRTBB0Cp0KOHVG7Q8cWqd0Z21H6w9oO4OKsTQEFEYQY6lxUBghwYJFFhRDCEn4kAAhIcmGfGw2995z3/4IdCju0gjtjM+vO+e+H89933Of8x7guwo/IAC/AED3EIYAvxiOdW+4GxIj8olmRHtWT3XUHHMuuXZdzc1MMTdu/Lz20g95qiWIoD6SoFMx1XKcgvrauUUpzRd5TZJLb5n+YHjHUxXBwRsmfNP2G6VhAMGM0oghleZQn5j/zzP0DnOpM4igThTd51byzKAggjrzDMeBY+aWrh6xaFAXZ5NC2UM3Ppj/wyFWIADMDFqU/VBHrofCp1rpzx+c/2QDiOFjn5qGNO5ABwHAzec6qjWYgXlZRU9/f6z5/LmrnNGuaqMOnzncFy15TAJ++AX8AVRvKBo/x2t8FZE2q8et40KHsa/yK9tKYP/F6Lx9qU/kYdP4DFHW2mtDvCUij58TBXPbZx8DgAAC5u0earQwDWigXQHIhR5O13VhbemyyGRfEeedOzjvWedg8OjXCzdrvQMHIgbHKwqgChmyu90FUzLDq8akKZ6+/BLj4u4jnJcwZElyKZkBBI764FMBjIzATRiSDIuVICK9NOknP1YuncuT6psvpk7wpK61GUNrrTe8iRT0kwM28xq0+Wtk/myv+tnWGqmmCMA0JDDcpmg5ohLwYhcDwPku7erE7ymGQlBP1dSxmpGt9BsW/sPTIQmdRWQIrAhAMlNiYsh88TUS4SEoZz75HNA0IVQrznWbV4ZjeqMSuNMmBOC1PO5Na7SqYtyBk7r5m2UsVj5mx5WrDLeL4LIDzIAhgfYQI8tDqK4ZxAt/UXlytiCnnbuqTmrjgNibMNYvxT74FKBBu9zFW89fEfjtCou58jEHLrQCm3ZL1J81EdEAXQL1ZyV2fqyjqYnx6AIX1j8nZH9YRctlcztwuK8c5Uq05DFbcGsdLrTj+ktPCjy+hLBvn4YTjTrOXlKxYQdQ8StCfByw8V1GOEKQMoLsJgVlD9vgsgisfAX9IKCDO2KqYqwK0EGulWBf4qICWrt4HmPrjojYdigJDeqDyL3fCbsVONnCaL4IAIz8PBtOyQew+8t0vFo1qEz/gcnLF6jPgOdm1lGtEStX1EUffAoJcLZKK7xjrelb9hjG5PGqWDxzEKPSnGjpTUWS00B6ikCck3BfkonTHW6kZiajZEoYJTNVqtpjyIlZatwkm3yGh2NGzRW1BbUoNNmspaL7uOTrdpMjmklNrQLPlmuYZa1Fv1fF6++ryMkCFCshPk7BmpnXMCa9Df2aiqoawtVOpgSn5MwUUVi/yi8oEJAjrkA5Gkgo4EENnSkJRNO8AtWHGK/tZmgWB1SrBU+WCiiKCRgSS4sF4lwKDJsD2z4EKj5gTMoBRo8i6hngbuV3ARPwR90HUQnswi5TEBDqx/iIzhgcYsQ7CJ+eJLz8poHR6Yy6emD5q2n4+eZMvHdEwTgP8MctOqrrgDg7QRsCaTqjbwA5UvoFoshwzBYAYAKQ7jYPMsvpCU5SZk1mZKQC3mxCUysje7TA3w53wzCBJWWMMxcZBfkEu43R2g4kuYk6esDuOP4UOCgwLMPf0IKoFfDDL6QJ2J3YeaKJxPkrzKluIDIEXAsxwtdNzMjVsHpGD1ZMCeGhSRHoYRNdPYy2ToLDBlwLmWawQZBVFbuFqDViaUFMHWAGckaL3mOnzY68MI3SDfD8HxFt3ytx5IwV6phkHE2YCD3Rho/+fgrKlU5YIkNYOFug7gvmcETQ+Ta9a5zHbOXO2FIcE374BTPokWzfz8pyinlprs8IVRbzyyVzODe1lBP3BxlhZkSYE4428gRPGa+ZMYv7thXzU9MK9bKcYn44u/A5Zgj/8GwZFXea2wgAr55a7D7eLJtVwak7f6/AHQda9oKObjgQ/sUj0Ox2JL9RDUtnD/4asMDtIl78a5PCgzzgyZITqhsPtfkBEYhyFAN3Hq+4HOVKRXD/9YwU/pPbRV9W1jB90WhIT148BjQB57pKuJ9/HZHuCDK9blxuk/KtvSCblU8kuXl9deOhtnKUK7GS/7cK/Pv9ntWljvc+G5hRfwEfeT2spmRY8e6jqzD2/b0wDROtKxajdPsbrLX10unLKlKSZElpUX9dtCH02xK41Y4L0wsrWBH3o6t/ztVfPmF2Ll2gkGki4cN/SM9LmxW4HR+zNDvqOg/+FBz9+L0dI7ow+IcJiMwk6X8gU64Ju+PXZ7zytkKNrboRCmuedRXKgN1RNX2iXJuRzOvAIH8M5fufgJlptmtm1fTcZTwtfznPds2sqaxcYf+/JbwFhOGqCRKEOa5Zm3xxBW8fYP9NPbnXq9y3InKDym1rdxXo7nEzPWMEG+47iX8BYMoY/w4QWDIAAAAASUVORK5CYII="
        pixmap = QPixmap()
        pixmap.loadFromData(QByteArray(base64.b64decode(ICON_B64)))
        return QIcon(pixmap)
    except:
        return QIcon()

def calisma_dizini_bul():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# ─────────────────────────────────────────
#  ÖZEL BAŞLIK ÇUBUĞU
# ─────────────────────────────────────────
class OzelBaslikCubugu(QWidget):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(15, 5, 10, 5)
        self.layout.setSpacing(10)
        self.setStyleSheet("background-color: #11111B; border-bottom: 1px solid #313244;")

        self.ikon = QLabel("☁️")
        self.ikon.setStyleSheet("border: none; font-size: 16px;")

        self.title_label = QLabel("StockFlow · Et & Tavuk Takip Sistemi")
        self.title_label.setStyleSheet("color: #CDD6F4; font-weight: bold; font-size: 13px; border: none;")

        btn_stil = "QPushButton { background-color: transparent; color: #A6ADC8; font-weight: bold; font-size: 16px; border: none; padding: 5px 10px; border-radius: 5px;} QPushButton:hover { background-color: #313244; color: #CDD6F4; }"
        kapat_btn_stil = "QPushButton { background-color: transparent; color: #A6ADC8; font-weight: bold; font-size: 16px; border: none; padding: 5px 10px; border-radius: 5px;} QPushButton:hover { background-color: #F38BA8; color: #11111B; }"

        self.min_btn = QPushButton("🗕")
        self.min_btn.setStyleSheet(btn_stil)
        self.min_btn.clicked.connect(self.parent.showMinimized)

        self.max_btn = QPushButton("🗖")
        self.max_btn.setStyleSheet(btn_stil)
        self.max_btn.clicked.connect(self.buyut_kucult)

        self.close_btn = QPushButton("✕")
        self.close_btn.setStyleSheet(kapat_btn_stil)
        self.close_btn.clicked.connect(self.parent.close)

        self.layout.addWidget(self.ikon)
        self.layout.addWidget(self.title_label)
        self.layout.addStretch()
        self.layout.addWidget(self.min_btn)
        self.layout.addWidget(self.max_btn)
        self.layout.addWidget(self.close_btn)

        self.initial_pos = None

    def buyut_kucult(self):
        if self.parent.isMaximized(): self.parent.showNormal()
        else: self.parent.showMaximized()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            # Maximize ise önce normal boyuta al, sonra sürüklemeye izin ver
            if self.parent.isMaximized():
                self.parent.showNormal()
                QApplication.processEvents()
            self.initial_pos = event.globalPosition().toPoint() - self.parent.pos()

    def mouseMoveEvent(self, event):
        if self.initial_pos is None: return
        global_pos = event.globalPosition().toPoint()
        self.parent.move(global_pos - self.initial_pos)

        # Snap önizleme
        ekran = self.parent.screen().availableGeometry()
        x, y = global_pos.x(), global_pos.y()

        if y <= 5:
            self.setStyleSheet("background-color: #1E3A5F; border-bottom: 2px solid #89B4FA;")
        elif x <= 5:
            self.setStyleSheet("background-color: #1A3A1A; border-bottom: 2px solid #A6E3A1;")
        elif x >= ekran.right() - 5:
            self.setStyleSheet("background-color: #1A3A1A; border-bottom: 2px solid #A6E3A1;")
        else:
            self.setStyleSheet("background-color: #11111B; border-bottom: 1px solid #313244;")

    def mouseReleaseEvent(self, event):
        if self.initial_pos is None: return
        global_pos = event.globalPosition().toPoint()
        ekran = self.parent.screen().availableGeometry()
        x, y = global_pos.x(), global_pos.y()

        if y <= 5:
            self.parent.showMaximized()
        elif x <= 5:
            self.parent.showNormal()
            QApplication.processEvents()
            self.parent.setGeometry(ekran.x(), ekran.y(), ekran.width() // 2, ekran.height())
        elif x >= ekran.right() - 5:
            self.parent.showNormal()
            QApplication.processEvents()
            self.parent.setGeometry(ekran.x() + ekran.width() // 2, ekran.y(), ekran.width() // 2, ekran.height())

        self.initial_pos = None
        self.setStyleSheet("background-color: #11111B; border-bottom: 1px solid #313244;")

    def mouseDoubleClickEvent(self, event):
        self.buyut_kucult()


class SiralanabilirItem(QTableWidgetItem):
    def __init__(self, gosterilen_metin, gercek_deger):
        super().__init__(gosterilen_metin)
        self.gercek_deger = gercek_deger

    def __lt__(self, other):
        if self.gercek_deger == "-": return False
        if other.gercek_deger == "-": return True
        try: return self.gercek_deger < other.gercek_deger
        except TypeError: return str(self.gercek_deger) < str(other.gercek_deger)


class ComboBoxDelegate(QStyledItemDelegate):
    """Belirli sütunlarda QComboBox göster"""
    def __init__(self, secenekler, parent=None):
        super().__init__(parent)
        self.secenekler = secenekler

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self.secenekler)
        combo.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #FAB387; border-radius: 6px; padding: 3px;")
        return combo

    def setEditorData(self, editor, index):
        val = index.data() or ""
        idx = editor.findText(val)
        if idx >= 0:
            editor.setCurrentIndex(idx)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText())


class SpinBoxDelegate(QStyledItemDelegate):
    """kg sütunları için geniş QDoubleSpinBox"""
    def createEditor(self, parent, option, index):
        w = QDoubleSpinBox(parent)
        w.setRange(0, 99999)
        w.setDecimals(3)
        w.setSingleStep(0.001)
        w.setSuffix(" kg")
        w.setMinimumWidth(120)
        w.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #FAB387; border-radius: 6px; padding: 4px 8px; font-size: 13px;")
        return w

    def setEditorData(self, editor, index):
        val = index.data() or "0"
        try:
            editor.setValue(float(str(val).replace("kg","").replace(",",".").strip()))
        except:
            editor.setValue(0)

    def setModelData(self, editor, model, index):
        model.setData(index, f"{editor.value():.3f} kg".replace(".", ","))


class DateDelegate(QStyledItemDelegate):
    """Tarih sütunları için QDateEdit — ESC ile iptal"""
    def createEditor(self, parent, option, index):
        from PyQt6.QtWidgets import QWidget, QHBoxLayout
        kap = QWidget(parent)
        kap.setStyleSheet("background: transparent;")
        lay = QHBoxLayout(kap)
        lay.setContentsMargins(0,0,0,0)
        lay.setSpacing(2)
        w = QDateEdit(kap)
        w.setCalendarPopup(True)
        w.setDisplayFormat("dd.MM.yyyy")
        w.setMinimumWidth(100)
        w.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #FAB387; border-radius: 6px; padding: 3px; font-size: 13px;")
        sil_btn = QPushButton("✕", kap)
        sil_btn.setFixedWidth(22)
        sil_btn.setStyleSheet("background-color: #F38BA8; color: #11111B; border-radius: 4px; font-size: 11px; font-weight: bold; padding: 0;")
        sil_btn.setToolTip("Tarihi temizle")
        lay.addWidget(w)
        lay.addWidget(sil_btn)
        kap._dateEdit = w
        kap._orijinal = index.data() or "-"
        kap._iptal = False
        kap._sil = False
        def on_sil():
            kap._sil = True
            self.commitData.emit(kap)
            self.closeEditor.emit(kap, QStyledItemDelegate.EndEditHint.NoHint)
        sil_btn.clicked.connect(on_sil)
        return kap

    def setEditorData(self, editor, index):
        val = index.data() or ""
        editor._orijinal = val
        w = getattr(editor, "_dateEdit", editor)
        if val and val != "-":
            try:
                d, m, y = val.split(".")
                w.setDate(QDate(int(y), int(m), int(d)))
            except:
                w.setDate(QDate.currentDate())
        else:
            w.setDate(QDate.currentDate())

    def setModelData(self, editor, model, index):
        if getattr(editor, "_iptal", False):
            model.setData(index, editor._orijinal)
        elif getattr(editor, "_sil", False):
            model.setData(index, "-")
        else:
            w = getattr(editor, "_dateEdit", editor)
            model.setData(index, w.date().toString("dd.MM.yyyy"))

    def eventFilter(self, editor, event):
        from PyQt6.QtCore import QEvent
        if event.type() == QEvent.Type.KeyPress:
            if event.key() == Qt.Key.Key_Escape:
                editor._iptal = True
                self.closeEditor.emit(editor, QStyledItemDelegate.EndEditHint.RevertModelCache)
                return True
        return super().eventFilter(editor, event)


class DynamicComboDelegate(QStyledItemDelegate):
    """Tablodaki mevcut değerlerden dinamik ComboBox (işletme gibi)"""
    def __init__(self, col_idx, extra=None, parent=None):
        super().__init__(parent)
        self.col_idx = col_idx
        self.extra = extra or []

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.setEditable(True)
        # Tablodan benzersiz değerleri topla
        tablo = index.model()
        degerler = set()
        for r in range(tablo.rowCount()):
            idx2 = tablo.index(r, self.col_idx)
            val = tablo.data(idx2) or ""
            if val and val != "-":
                degerler.add(val)
        secenekler = ["-"] + sorted(degerler) + [e for e in self.extra if e not in degerler]
        combo.addItems(secenekler)
        combo.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #FAB387; border-radius: 6px; padding: 3px;")
        return combo

    def setEditorData(self, editor, index):
        val = index.data() or "-"
        idx = editor.findText(val)
        if idx >= 0:
            editor.setCurrentIndex(idx)
        else:
            editor.setEditText(val)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText())


class GelistirilmisTablo(QTableWidget):
    def __init__(self, ana_pencere=None):
        super().__init__()
        self.ana_pencere = ana_pencere
        self.suanki_font_boyutu = 12

    def wheelEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            delta = 1 if event.angleDelta().y() > 0 else -1
            if self.ana_pencere:
                self.ana_pencere.font_boyutu_degistir(delta)
            else:
                self.suanki_font_boyutu = max(8, min(26, self.suanki_font_boyutu + delta))
                self._fontu_uygula()
            event.accept()
        else:
            super().wheelEvent(event)

    def _fontu_uygula(self):
        self.setFont(QFont("Arial", self.suanki_font_boyutu))
        baslik_font = self.horizontalHeader().font()
        baslik_font.setPointSize(max(8, self.suanki_font_boyutu - 2))
        baslik_font.setBold(True)
        self.horizontalHeader().setFont(baslik_font)
        # Satır yüksekliği: font boyutuna göre adaptif
        satir_yuksekligi = max(26, int(self.suanki_font_boyutu * 2.2))
        self.verticalHeader().setDefaultSectionSize(satir_yuksekligi)
        self.resizeColumnsToContents()

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.StandardKey.Copy): self.kopyala(); return
        elif event.matches(QKeySequence.StandardKey.Paste): self.yapistir(); return
        elif event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if self.state() != QAbstractItemView.State.EditingState:
                suanki_hucre = self.currentItem()
                if suanki_hucre and (suanki_hucre.flags() & Qt.ItemFlag.ItemIsEditable):
                    self.editItem(suanki_hucre); return
        super().keyPressEvent(event)

    def kopyala(self):
        secili_araliklar = self.selectedRanges()
        if not secili_araliklar: return
        metin = ""
        aralik = secili_araliklar[0]
        for row in range(aralik.topRow(), aralik.bottomRow() + 1):
            satir_verisi = []
            for col in range(aralik.leftColumn(), aralik.rightColumn() + 1):
                item = self.item(row, col)
                satir_verisi.append(item.text().replace(" kg", "").replace("(KRİTİK)", "").strip() if item else "")
            metin += "\t".join(satir_verisi) + "\n"
        QApplication.clipboard().setText(metin)

    def yapistir(self):
        metin = QApplication.clipboard().text()
        if not metin: return
        suanki_hucre = self.currentItem()
        if not suanki_hucre: return
        baslangic_satir = suanki_hucre.row()
        baslangic_sutun = suanki_hucre.column()
        satirlar = metin.split('\n')
        self.blockSignals(True)
        degisiklik_yapildi = False
        for i, satir in enumerate(satirlar):
            if not satir.strip(): continue
            for j, deger in enumerate(satir.split('\t')):
                hedef_satir = baslangic_satir + i
                hedef_sutun = baslangic_sutun + j
                if hedef_satir < self.rowCount() and hedef_sutun < self.columnCount():
                    item = self.item(hedef_satir, hedef_sutun)
                    if item and (item.flags() & Qt.ItemFlag.ItemIsEditable):
                        if self.ana_pencere:
                            self.ana_pencere.hizli_hucre_guncelle(hedef_satir, hedef_sutun, deger)
                            degisiklik_yapildi = True
        self.blockSignals(False)
        if degisiklik_yapildi and self.ana_pencere:
            self.ana_pencere.verileri_yukle()

    def contextMenuEvent(self, event):
        if not self.ana_pencere: return
        item = self.itemAt(event.pos())
        if not item: return
        satir = item.row()

        # Transfer durumunu kontrol et
        id_col = self.columnCount() - 1
        id_item = self.item(satir, id_col)
        if not id_item: return
        urun_id = int(id_item.text())

        # Transfer var mı?
        transfer_col = None
        for col in range(self.columnCount()):
            h = self.horizontalHeaderItem(col)
            if h and h.text() in ["Transfer (kg)", "Transfer Yön"]:
                transfer_col = col - (0 if "kg" in (h.text() or "") else 2)
                break
        # Transfer miktar sütunu 12
        tr_item = self.item(satir, 12)
        transfer_var = False
        if tr_item:
            try:
                val = float(tr_item.text().replace("kg","").replace(",",".").strip())
                transfer_var = val > 0
            except: pass

        from PyQt6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { background-color: #1E1E2E; color: #CDD6F4; border: 1px solid #45475A;
                    border-radius: 8px; padding: 4px; font-size: 13px; }
            QMenu::item { padding: 8px 20px; border-radius: 6px; }
            QMenu::item:selected { background-color: #313244; }
            QMenu::separator { height: 1px; background: #313244; margin: 4px 8px; }
        """)

        akt_duzenle = menu.addAction("✏️  Düzenle / Modal Aç")
        akt_kopyala = menu.addAction("📋  Kopyala")
        menu.addSeparator()
        if transfer_var:
            akt_transfer = menu.addAction("🔄  Transfer İptal Et")
        else:
            akt_transfer = menu.addAction("🔄  Transfer Yap")
        menu.addSeparator()
        akt_sil = menu.addAction("🗑  Sil")

        secilen = menu.exec(event.globalPos())

        if secilen == akt_duzenle:
            if self.ana_pencere:
                self.ana_pencere._duzenle_dialog_ac(urun_id)

        elif secilen == akt_kopyala:
            self.kopyala()

        elif secilen == akt_transfer:
            if transfer_var:
                # Transfer iptal
                cevap = QMessageBox.question(self.ana_pencere, "Transfer İptal",
                    "Bu ürünün transferini iptal etmek istiyor musunuz?\n"
                    "Transfer bilgileri silinecek, kalan miktar güncellenir.",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if cevap == QMessageBox.StandardButton.Yes:
                    # Mevcut ilk ve diğer düşümler
                    try:
                        ilk = float((self.item(satir,7) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                        kuvet = float((self.item(satir,4) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                        takoz = float((self.item(satir,6) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                        zayi = float((self.item(satir,10) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                        yeni_kalan = max(0.0, ilk - kuvet - takoz - zayi)
                        self.ana_pencere.api.urun_guncelle(urun_id, {
                            "transfer_miktar": 0.0, "transfer_tarihi": "-",
                            "transfer_yon": "-", "transfer_isletme": "-",
                            "kalan_miktar": yeni_kalan
                        })
                        self.ana_pencere.verileri_yukle()
                    except Exception as e:
                        QMessageBox.critical(self.ana_pencere, "Hata", str(e))
            else:
                # Transfer diyalogu
                self.ana_pencere._transfer_dialog_ac(urun_id, satir, self)

        elif secilen == akt_sil:
            cevap = QMessageBox.question(self.ana_pencere, "Sil",
                "Bu kaydı silmek istediğinize emin misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if cevap == QMessageBox.StandardButton.Yes:
                self.ana_pencere.api.urun_sil(urun_id)
                self.ana_pencere.verileri_yukle()


class SutunFiltreDialog(QDialog):
    def __init__(self, sutun_adi, benzersiz_degerler, aktif_secimler, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Filtrele: {sutun_adi}")
        self.resize(300, 350)
        self.setWindowFlags(Qt.WindowType.Popup)
        layout = QVBoxLayout(self)
        self.arama = QLineEdit()
        self.arama.setPlaceholderText("Ara...")
        self.arama.textChanged.connect(self.filtre_ici_arama)
        layout.addWidget(self.arama)
        self.liste = QListWidget()
        self.liste.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 1px solid #45475A; border-radius: 8px; padding: 5px;")
        layout.addWidget(self.liste)
        self.itemler = []
        for val in sorted(benzersiz_degerler, key=lambda x: (x == "-", x)):
            item = QListWidgetItem(val)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(Qt.CheckState.Checked if not aktif_secimler or val in aktif_secimler else Qt.CheckState.Unchecked)
            self.liste.addItem(item)
            self.itemler.append(item)
        btn_layout = QHBoxLayout()
        btn_all = QPushButton("Tümünü Seç")
        btn_all.clicked.connect(lambda: self.secim_yap(Qt.CheckState.Checked))
        btn_none = QPushButton("Temizle")
        btn_none.clicked.connect(lambda: self.secim_yap(Qt.CheckState.Unchecked))
        btn_layout.addWidget(btn_all)
        btn_layout.addWidget(btn_none)
        layout.addLayout(btn_layout)
        btn_uygula = QPushButton("Filtreyi Uygula")
        btn_uygula.setStyleSheet("background-color: #A6E3A1; color: #11111B;")
        btn_uygula.clicked.connect(self.accept)
        layout.addWidget(btn_uygula)

    def filtre_ici_arama(self, text):
        text = text.lower()
        for i in range(self.liste.count()):
            self.liste.item(i).setHidden(text not in self.liste.item(i).text().lower())

    def secim_yap(self, durum):
        for i in range(self.liste.count()):
            if not self.liste.item(i).isHidden():
                self.liste.item(i).setCheckState(durum)

    def secilenleri_al(self):
        return [self.liste.item(i).text() for i in range(self.liste.count()) if self.liste.item(i).checkState() == Qt.CheckState.Checked]


class UrunEkleDialog(QDialog):
    def __init__(self, parent=None, varsayilan_kategori=None):
        super().__init__(parent)
        self.setWindowTitle("Yeni Kayıt")
        self.resize(450, 250)
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.barkod_input = QLineEdit()
        self.barkod_input.setPlaceholderText("Barkodu okutun...")
        self.barkod_input.textChanged.connect(self.barkoddan_kilo_hesapla)
        self.barkod_oku_btn = QPushButton("📷 Görselden Oku")
        self.barkod_oku_btn.setStyleSheet("QPushButton { background-color: #89B4FA; color: #11111B; border-radius: 10px; padding: 8px; font-weight: bold;} QPushButton:hover { background-color: #B4BEFE; }")
        self.barkod_oku_btn.clicked.connect(self.gorselden_barkod_oku)
        barkod_layout = QHBoxLayout()
        barkod_layout.addWidget(self.barkod_input)
        barkod_layout.addWidget(self.barkod_oku_btn)
        self.kategori_input = QComboBox()
        self.kategori_input.addItems(["Et", "Tavuk"])
        if varsayilan_kategori:
            self.kategori_input.setCurrentText(varsayilan_kategori)
            self.kategori_input.setEnabled(False)  # Sekmeden açılınca kategori sabitle
        self.kategori_input.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 12px; padding: 5px;")
        ortak_stil = "background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 12px; padding: 5px;"
        self.gelis_tarihi = QDateEdit()
        self.gelis_tarihi.setCalendarPopup(True)
        self.gelis_tarihi.setDate(QDate.currentDate())
        self.gelis_tarihi.setStyleSheet(ortak_stil)
        self.ilk_miktar_input = QDoubleSpinBox()
        self.ilk_miktar_input.setRange(0, 100000.00)
        self.ilk_miktar_input.setDecimals(3)
        self.ilk_miktar_input.setSuffix(" kg")
        self.ilk_miktar_input.setStyleSheet(ortak_stil)
        form_layout.addRow("Barkod Numarası:", barkod_layout)
        form_layout.addRow("Kategori:", self.kategori_input)
        form_layout.addRow("Geliş Tarihi:", self.gelis_tarihi)
        form_layout.addRow("İlk Miktar:", self.ilk_miktar_input)
        layout.addLayout(form_layout)
        self.kaydet_btn = QPushButton("Kaydet")
        self.kaydet_btn.clicked.connect(self.accept)
        layout.addWidget(self.kaydet_btn)

    def barkoddan_kilo_hesapla(self, barkod):
        if len(barkod) >= 12:
            try:
                hesaplanan_kilo = float(f"{barkod[8:10]}.{barkod[10:12]}")
                if len(barkod) >= 13 and int(barkod[12]) > 5:
                    hesaplanan_kilo += 0.01
                self.ilk_miktar_input.setValue(hesaplanan_kilo)
            except ValueError:
                pass

    def gorselden_barkod_oku(self):
        dosya_yolu, _ = QFileDialog.getOpenFileName(self, "Etiket Görseli Seç", "", "Resim Dosyaları (*.png *.jpg *.jpeg *.bmp)")
        if dosya_yolu:
            try:
                barkodlar = pyzbar_decode(Image.open(dosya_yolu))
                if barkodlar: self.barkod_input.setText(barkodlar[0].data.decode('utf-8'))
                else: QMessageBox.warning(self, "Hata", "Görselde barkod bulunamadı.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))

    def verileri_al(self):
        return (self.barkod_input.text(), self.kategori_input.currentText(),
                self.gelis_tarihi.date().toString("dd.MM.yyyy"), "-", "-",
                self.ilk_miktar_input.value(), self.ilk_miktar_input.value())


# ─────────────────────────────────────────
#  ANA SİSTEM
# ─────────────────────────────────────────
class StokSistemi(QMainWindow):
    def __init__(self, api: ApiIstemci):
        super().__init__()
        self.api = api  # ← Giriş ekranından gelen oturumlu API istemcisi

        self.resize(1400, 800)  # geçici, showEvent'te güncellenir

        self.aktif_filtreler = {}
        self.veri_klasoru = calisma_dizini_bul()
        self.ilk_acilis = True

        ayarlar_yolu = os.path.join(self.veri_klasoru, "ayarlar.ini")
        self.ayarlar = QSettings(ayarlar_yolu, QSettings.Format.IniFormat)

        self.aylar = {"01": "Ocak", "02": "Şubat", "03": "Mart", "04": "Nisan",
                      "05": "Mayıs", "06": "Haziran", "07": "Temmuz", "08": "Ağustos",
                      "09": "Eylül", "10": "Ekim", "11": "Kasım", "12": "Aralık"}

        # Global font boyutu — ayarlardan yükle
        self.global_font_boyutu = int(self.ayarlar.value("font_boyutu", 12))

        self.setWindowTitle("StockFlow · Et & Tavuk Takip Sistemi")
        self.tema_uygula()
        self.init_ui()
        self.verileri_yukle()
        # Güncelleme kontrolü — arka planda
        QTimer.singleShot(3000, self.guncelleme_kontrol_et)
        # Ctrl+scroll global olarak yakala
        QApplication.instance().installEventFilter(self)

    def _yerel_yedek_kaydet(self, veriler):
        """Verileri yerel JSON dosyasına yedekle"""
        try:
            import json
            yedek_yolu = os.path.join(self.veri_klasoru, YEDEK_DOSYA)
            yedek = {
                "tarih": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "versiyon": UYGULAMA_VERSIYON,
                "kayit_sayisi": len(veriler),
                "veriler": veriler
            }
            with open(yedek_yolu, 'w', encoding='utf-8') as f:
                json.dump(yedek, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            print(f"Yedekleme hatası: {e}")

    def _yerel_yedekten_yukle(self):
        """Cloudflare erişilemiyorsa yerel yedekten yükle"""
        try:
            import json
            yedek_yolu = os.path.join(self.veri_klasoru, YEDEK_DOSYA)
            if not os.path.exists(yedek_yolu):
                return None
            with open(yedek_yolu, 'r', encoding='utf-8') as f:
                yedek = json.load(f)
            return yedek
        except Exception:
            return None

    def guncelleme_kontrol_et(self):
        self.gkontrolcu = GuncellemKontrolcu()
        self.gkontrolcu.guncelleme_var.connect(self._guncelleme_bildir)
        self.gkontrolcu.start()

    def _guncelleme_bildir(self, yeni_versiyon):
        cevap = QMessageBox.question(self, "Güncelleme Mevcut",
            f"Yeni sürüm: {yeni_versiyon}\nMevcut sürüm: {UYGULAMA_VERSIYON}\n\n"
            "İndirme sayfası tarayıcıda açılacak.\n"
            "İndirilen exe'yi çalıştırarak güncelleme yapabilirsiniz.",
            QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
        if cevap == QMessageBox.StandardButton.Ok:
            import webbrowser
            webbrowser.open(GITHUB_RELEASE_URL)

    def _guncellemeyi_indir(self, yeni_versiyon):
        import webbrowser, tempfile, subprocess
        # Önce direkt indirmeyi dene
        try:
            self.lbl_durum.setText("⬇️ İndiriliyor...")
            QApplication.processEvents()
            # GitHub redirect'i takip ederek indir
            oturum = requests.Session()
            oturum.headers.update({"User-Agent": "StockFlow-Updater/1.0"})
            r = oturum.get(GITHUB_RELEASE_URL, stream=True, timeout=60, allow_redirects=True)
            if r.status_code == 200:
                # Masaüstüne indir
                masaustu = os.path.join(os.path.expanduser("~"), "Desktop")
                hedef = os.path.join(masaustu, "StockFlow_v15_yeni.exe")
                with open(hedef, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                self.lbl_durum.setText("✅ İndirildi!")
                QMessageBox.information(self, "Güncelleme Hazır",
                    f"Yeni exe masaüstüne indirildi:\nStockFlow_v15_yeni.exe\n\n"
                    "Uygulamayı kapatıp yeni exe'yi çalıştırın.")
            else:
                raise Exception(f"HTTP {r.status_code}")
        except Exception as e:
            self.lbl_durum.setText("")
            webbrowser.open("https://github.com/Mirza1293/vasak-erp-backend/releases/latest")

    def _guncelleme_tamamlandi(self, dosya_yolu):
        self.lbl_durum.setText("✅ Güncelleme hazır!")
        try:
            import subprocess, shutil, tempfile
            hedef_klasor = calisma_dizini_bul()
            hedef_exe = os.path.join(hedef_klasor, "StockFlow_v15.exe")
            # Bat ile: bekle → kopyala → başlat
            bat_icerik = (
                "@echo off\r\n"
                "timeout /t 2 /nobreak >nul\r\n"
                f'copy /y "{dosya_yolu}" "{hedef_exe}"\r\n'
                f'start "" "{hedef_exe}"\r\n'
                "del %~f0\r\n"
            )
            bat_yol = os.path.join(tempfile.gettempdir(), "sf_guncelle.bat")
            with open(bat_yol, 'w', encoding='cp1254') as f:
                f.write(bat_icerik)
            QMessageBox.information(self, "Güncelleme Hazır",
                "Güncelleme tamamlandı!\nUygulama yeniden başlatılacak.")
            subprocess.Popen(f'cmd /c "{bat_yol}"', shell=True)
            QApplication.quit()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Güncelleme hatası: {e}")

    def tema_degistir(self, tema_adi):
        global AKTIF_TEMA
        AKTIF_TEMA = tema_adi
        self.tema_uygula()
        # Ayara kaydet
        self.ayarlar.setValue("tema", tema_adi)
        self.ayarlar.sync()

    def closeEvent(self, event):
        et_col = self.table_et.horizontalHeader().sortIndicatorSection()
        et_order = self.table_et.horizontalHeader().sortIndicatorOrder().value
        tavuk_col = self.table_tavuk.horizontalHeader().sortIndicatorSection()
        tavuk_order = self.table_tavuk.horizontalHeader().sortIndicatorOrder().value
        self.ayarlar.setValue("et_sort_col", et_col if et_col >= 0 else 0)
        self.ayarlar.setValue("et_sort_order", et_order)
        self.ayarlar.setValue("tavuk_sort_col", tavuk_col if tavuk_col >= 0 else 0)
        self.ayarlar.setValue("tavuk_sort_order", tavuk_order)
        self.ayarlar.setValue("font_boyutu", self.global_font_boyutu)
        self.ayarlar.sync()
        super().closeEvent(event)

    def _sutun_genisliklerini_ayarla(self):
        for tbl in [self.table_et, self.table_tavuk]:
            tbl.resizeColumnsToContents()
            header = tbl.horizontalHeader()
            # Header kendi fontunu kullanarak başlık genişliklerini hesapla
            hfm = header.fontMetrics()
            for i in range(tbl.columnCount()):
                item = tbl.horizontalHeaderItem(i)
                if item:
                    # Her satır için en uzun metni hesapla
                    baslik_gen = hfm.horizontalAdvance(item.text()) + 30
                    if tbl.columnWidth(i) < baslik_gen:
                        tbl.setColumnWidth(i, baslik_gen)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        if event.type() == QEvent.Type.Wheel:
            if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
                delta = 1 if event.angleDelta().y() > 0 else -1
                self.font_boyutu_degistir(delta)
                return True
        return super().eventFilter(obj, event)

    def font_boyutu_degistir(self, delta):
        self.global_font_boyutu = max(8, min(26, self.global_font_boyutu + delta))
        # Tüm tablolar ölçeklendirilsin
        tum_tablolar = [
            self.table_et, self.table_tavuk,
            self.table_gunluk, self.table_haftalik,
            self.table_aylik, self.table_gelen,
            self.table_transfer
        ]
        for tbl in tum_tablolar:
            tbl.suanki_font_boyutu = self.global_font_boyutu
            tbl._fontu_uygula()
        # Genel uygulama fontu da güncelle
        QApplication.instance().setFont(QFont("Arial", self.global_font_boyutu))
        self.lbl_durum.setText(f"🔤 Punto: {self.global_font_boyutu}")
        QTimer.singleShot(1500, lambda: self.lbl_durum.setText(""))

    def showEvent(self, event):
        super().showEvent(event)
        ekran = self.screen().availableGeometry()
        dpi = self.screen().logicalDotsPerInch()

        # DPI'a göre pencere boyutu oranı
        if dpi <= 96:    oran = 0.88
        elif dpi <= 120: oran = 0.90
        else:            oran = 0.92

        genislik = int(ekran.width() * oran)
        yukseklik = int(ekran.height() * oran)
        self.resize(genislik, yukseklik)
        self.move(
            ekran.x() + (ekran.width() - genislik) // 2,
            ekran.y() + (ekran.height() - yukseklik) // 2
        )
        # Kaydedilmiş font boyutunu uygula
        QTimer.singleShot(400, lambda: self.font_boyutu_degistir(0))

    def resizeEvent(self, event):
        super().resizeEvent(event)

    def tema_uygula(self):
        self.setStyleSheet("""
            QWidget#ana_widget { background-color: #181825; }
            QMainWindow, QDialog, QMessageBox { background-color: #181825; color: #CDD6F4; }
            QLabel { color: #CDD6F4; font-weight: bold; }
            QCheckBox { color: #CDD6F4; font-weight: bold; }
            QLineEdit { background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 12px; padding: 10px; font-size: 13px; }
            QLineEdit:focus { border: 2px solid #FAB387; }
            QPushButton { background-color: #FAB387; color: #11111B; border-radius: 12px; padding: 10px 15px; font-weight: bold; font-size: 13px; }
            QPushButton:hover { background-color: #F5E0DC; }
            QPushButton#btn_sil { background-color: #F38BA8; }
            QPushButton#btn_sil:hover { background-color: #eba0ac; }
            QPushButton#btn_excel { background-color: #A6E3A1; }
            QPushButton#btn_excel:hover { background-color: #94E2D5; }
            QPushButton#btn_import { background-color: #CBA6F7; }
            QPushButton#btn_import:hover { background-color: #B4BEFE; }
            QPushButton#btn_guvenlik { background-color: #89B4FA; }
            QPushButton#btn_guvenlik:hover { background-color: #B4BEFE; }
            QTabWidget::pane { border: none; }
            QTabBar::tab { background-color: #1E1E2E; color: #A6ADC8; padding: 12px 25px; margin-right: 5px; border-top-left-radius: 15px; border-top-right-radius: 15px; font-weight: bold; font-size: 14px; }
            QTabBar::tab:selected { background-color: #313244; color: #CDD6F4; }
            QMainWindow { background-color: #181825; }
            QMenuBar { background-color: #11111B; color: #CDD6F4; }
            QStatusBar { background-color: #11111B; color: #CDD6F4; }
            QScrollBar:vertical { background: #1E1E2E; width: 8px; border-radius: 4px; }
            QScrollBar::handle:vertical { background: #45475A; border-radius: 4px; }
            QScrollBar::handle:vertical:hover { background: #585B70; }
            QTableWidget { background-color: #1E1E2E; color: #CDD6F4; border: none; border-radius: 15px; gridline-color: transparent; selection-background-color: #45475A; outline: none; }
            QTableWidget::item:selected { border: 2px solid #FAB387; background-color: #45475A; }
            QHeaderView::section { background-color: #1E1E2E; color: #A6ADC8; padding: 12px; border: 1px solid #313244; font-weight: bold; text-transform: uppercase; font-size: 11px;}
            QHeaderView::section:hover { background-color: #313244; }
            QLabel#kard_et { background-color: #F38BA8; color: #11111B; border-radius: 20px; padding: 30px; }
            QLabel#kard_tavuk { background-color: #F9E2AF; color: #11111B; border-radius: 20px; padding: 30px; }
            QLabel#lbl_secim { background-color: #313244; color: #A6E3A1; padding: 10px; border-radius: 10px; font-weight: bold; font-size: 14px; }
        """)

    def init_ui(self):
        self.ana_widget = QWidget()
        self.ana_widget.setObjectName("ana_widget")
        self.setCentralWidget(self.ana_widget)
        main_layout = QVBoxLayout(self.ana_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        # Native başlık çubuğu kullanılıyor — özel başlık çubuğu devre dışı
        icerik_widget = QWidget()
        layout = QVBoxLayout(icerik_widget)
        layout.setContentsMargins(20, 15, 20, 15)
        layout.setSpacing(15)
        main_layout.addWidget(icerik_widget)

        top_bar = QHBoxLayout()
        self.arama_kutusu = QLineEdit()
        self.arama_kutusu.setPlaceholderText("🔍 Filtrele...")
        self.arama_kutusu.textChanged.connect(self.tablolari_filtrele)
        self.arama_kutusu.setMinimumWidth(200)
        self.import_btn = QPushButton("Excel'den Al")
        self.import_btn.setObjectName("btn_import")
        self.import_btn.clicked.connect(self.excelden_iceri_aktar)
        self.excel_btn = QPushButton("Tümünü Excel'e Aktar")
        self.excel_btn.setObjectName("btn_excel")
        self.excel_btn.clicked.connect(self.excel_disa_aktar)
        self.delete_btn = QPushButton("Sil")
        self.delete_btn.setObjectName("btn_sil")
        self.delete_btn.clicked.connect(self.urun_sil)
        self.add_btn = QPushButton("+ Yeni Kayıt")
        self.add_btn.clicked.connect(self.yeni_urun_penceresi_ac)
        top_bar.addWidget(self.arama_kutusu)
        top_bar.addStretch()
        # Tema seçici
        self.tema_combo = QComboBox()
        self.tema_combo.addItems(list(TEMALAR.keys()))
        self.tema_combo.setCurrentText(AKTIF_TEMA)
        self.tema_combo.setFixedWidth(100)
        self.tema_combo.currentTextChanged.connect(self.tema_degistir)
        top_bar.addWidget(QLabel("🎨"))
        top_bar.addWidget(self.tema_combo)
        top_bar.addWidget(self.import_btn)
        top_bar.addWidget(self.excel_btn)
        top_bar.addWidget(self.delete_btn)
        top_bar.addWidget(self.add_btn)
        layout.addLayout(top_bar)

        self.tabs = QTabWidget()
        self.tab_dashboard = QWidget()
        self.tab_et = QWidget()
        self.tab_tavuk = QWidget()
        self.tab_analiz = QWidget()
        self.tab_transfer = QWidget()
        self.tabs.addTab(self.tab_dashboard, "🏠 Ana Sayfa")
        self.tabs.addTab(self.tab_et, "🥩 Et Kayıtları")
        self.tabs.addTab(self.tab_tavuk, "🍗 Tavuk Kayıtları")
        self.tabs.addTab(self.tab_analiz, "📊 Analiz & Raporlar")
        self.tabs.addTab(self.tab_transfer, "🔄 Transfer")
        self.tab_kisayollar = QWidget()
        self.tabs.addTab(self.tab_kisayollar, "⌨️ Kısayollar")
        layout.addWidget(self.tabs)

        dash_layout = QHBoxLayout(self.tab_dashboard)
        self.lbl_et_toplam = QLabel("🥩 ET STOK DURUMU\n\nGiren: 0,00 kg\nKalan: 0,00 kg")
        self.lbl_et_toplam.setObjectName("kard_et")
        self.lbl_et_toplam.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_et_toplam.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        self.lbl_tavuk_toplam = QLabel("🍗 TAVUK STOK DURUMU\n\nGiren: 0,00 kg\nKalan: 0,00 kg")
        self.lbl_tavuk_toplam.setObjectName("kard_tavuk")
        self.lbl_tavuk_toplam.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_tavuk_toplam.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        dash_layout.addWidget(self.lbl_et_toplam)
        dash_layout.addWidget(self.lbl_tavuk_toplam)

        et_layout = QVBoxLayout(self.tab_et)
        self.table_et = self.tablo_olustur()
        et_layout.addWidget(self.table_et)
        tavuk_layout = QVBoxLayout(self.tab_tavuk)
        self.table_tavuk = self.tablo_olustur()
        tavuk_layout.addWidget(self.table_tavuk)

        analiz_layout = QVBoxLayout(self.tab_analiz)
        analiz_layout.setContentsMargins(8, 8, 8, 8)
        analiz_layout.setSpacing(8)

        # ── Özet kartları (her zaman görünür) ──
        ort_kapsayici = QHBoxLayout()
        ort_kapsayici.setSpacing(10)
        self.lbl_ort_7 = QLabel("⏳ Son 7 Gün\n\nEt: 0,00 kg/gün\nTavuk: 0,00 kg/gün")
        self.lbl_ort_7.setStyleSheet("background-color: #F9E2AF; color: #11111B; border-radius: 12px; padding: 12px; font-size: 14px; font-weight: bold;")
        self.lbl_ort_7.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_ort_30 = QLabel("📅 Son 30 Gün\n\nEt: 0,00 kg/gün\nTavuk: 0,00 kg/gün")
        self.lbl_ort_30.setStyleSheet("background-color: #89B4FA; color: #11111B; border-radius: 12px; padding: 12px; font-size: 14px; font-weight: bold;")
        self.lbl_ort_30.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ort_kapsayici.addWidget(self.lbl_ort_7)
        ort_kapsayici.addWidget(self.lbl_ort_30)
        analiz_layout.addLayout(ort_kapsayici)

        # ── Alt sekmeler ──
        self.analiz_tabs = QTabWidget()
        self.analiz_tabs.setStyleSheet("""
            QTabBar::tab { background-color: #313244; color: #A6ADC8; padding: 8px 18px;
                           margin-right: 3px; border-top-left-radius: 10px; border-top-right-radius: 10px;
                           font-weight: bold; font-size: 13px; }
            QTabBar::tab:selected { background-color: #45475A; color: #CDD6F4; }
            QTabWidget::pane { border: 1px solid #313244; border-radius: 8px; }
        """)

        # Alt sekme 1: Tüketim Analizi
        tab_tuketim = QWidget()
        tuketim_layout = QVBoxLayout(tab_tuketim)
        tuketim_layout.setContentsMargins(6, 6, 6, 6)
        tuketim_layout.setSpacing(6)
        tablolar_layout = QHBoxLayout()
        tablolar_layout.setSpacing(8)
        self.table_gunluk = self._analiz_tablosu(["Günlük Periyot", "Et (kg)", "Tavuk (kg)", "Toplam (kg)", "Zayi (kg)", "Paket Dağılımı"])
        self.table_haftalik = self._analiz_tablosu(["Haftalık Periyot", "Et (kg)", "Tavuk (kg)", "Toplam (kg)", "Zayi (kg)", "Paket Dağılımı"])
        self.table_aylik = self._analiz_tablosu(["Aylık Periyot", "Et (kg)", "Tavuk (kg)", "Toplam (kg)", "Zayi (kg)", "Paket Dağılımı"])
        tablolar_layout.addWidget(self.table_gunluk)
        tablolar_layout.addWidget(self.table_haftalik)
        tablolar_layout.addWidget(self.table_aylik)
        tuketim_layout.addLayout(tablolar_layout)
        self.analiz_tabs.addTab(tab_tuketim, "📉 Tüketim Analizi")

        # Alt sekme 2: Gelen Ürün Analizi
        tab_gelen = QWidget()
        gelen_layout = QVBoxLayout(tab_gelen)
        gelen_layout.setContentsMargins(6, 6, 6, 6)
        self.table_gelen = self._analiz_tablosu(["Geliş Tarihi", "Kategori", "4'lük\n(3-5)", "6'lık\n(5-7)", "8'lik\n(7-9)", "10'luk\n(9-11)", "12'lik\n(11-13)", "15'lik\n(14-17)", "20'lik\n(19-21)", "30'luk\n(28-32)", "40'lık\n(38-42)", "50'lik\n(48-52)", "Diğer", "Toplam (kg)"])
        gelen_layout.addWidget(self.table_gelen)
        self.analiz_tabs.addTab(tab_gelen, "📦 Gelen Ürün Analizi")

        # Alt sekme 3: Stok Durumu Özeti
        tab_stok = QWidget()
        stok_layout = QVBoxLayout(tab_stok)
        stok_layout.setContentsMargins(6, 6, 6, 6)
        self.table_stok_ozet = self._analiz_tablosu(["Kategori", "Toplam Giren (kg)", "Toplam Kalan (kg)", "Toplam Tüketilen (kg)", "Toplam Zayi (kg)", "Tüketim Oranı (%)"])
        stok_layout.addWidget(self.table_stok_ozet)
        self.analiz_tabs.addTab(tab_stok, "📊 Stok Özeti")

        # Alt sekme: Paket Raporu
        tab_paket = QWidget()
        paket_layout = QVBoxLayout(tab_paket)
        paket_layout.setContentsMargins(6, 6, 6, 6)
        self.table_paket_rapor = self._analiz_tablosu(["Tarih", "Et (Paket Dağılımı)", "Tavuk (Paket Dağılımı)"])
        paket_layout.addWidget(self.table_paket_rapor)
        self.analiz_tabs.addTab(tab_paket, "📦 Paket Raporu")

        analiz_layout.addWidget(self.analiz_tabs)

        # Excel rapor butonu
        self.btn_detayli_rapor = QPushButton("📄 Tüm Analizleri Excel'e Aktar")
        self.btn_detayli_rapor.setObjectName("btn_excel")
        self.btn_detayli_rapor.setStyleSheet("padding: 10px; font-size: 14px;")
        self.btn_detayli_rapor.clicked.connect(self.analiz_raporu_disa_aktar)
        analiz_layout.addWidget(self.btn_detayli_rapor)

        # ── TRANSFER SEKMESİ ──
        transfer_layout = QVBoxLayout(self.tab_transfer)
        transfer_layout.setContentsMargins(10, 10, 10, 10)

        # Üst bar: filtre butonları + yeni transfer butonu
        tr_top = QHBoxLayout()
        self.btn_tr_tumü = QPushButton("Tümü")
        self.btn_tr_cikis = QPushButton("→ Çıkış")
        self.btn_tr_giris = QPushButton("← Giriş")
        self.btn_tr_tumü.setStyleSheet("background-color: #FAB387; color: #11111B; border-radius: 8px; padding: 8px 14px; font-weight: bold;")
        self.btn_tr_cikis.setStyleSheet("background-color: #313244; color: #CDD6F4; border-radius: 8px; padding: 8px 14px;")
        self.btn_tr_giris.setStyleSheet("background-color: #313244; color: #CDD6F4; border-radius: 8px; padding: 8px 14px;")
        self.btn_tr_tumü.clicked.connect(lambda: self.transfer_filtrele("hepsi"))
        self.btn_tr_cikis.clicked.connect(lambda: self.transfer_filtrele("Çıkış"))
        self.btn_tr_giris.clicked.connect(lambda: self.transfer_filtrele("Giriş"))
        tr_top.addWidget(self.btn_tr_tumü)
        tr_top.addWidget(self.btn_tr_cikis)
        tr_top.addWidget(self.btn_tr_giris)
        tr_top.addStretch()
        self.btn_tr_iptal = QPushButton("🔄 Transferi İptal Et")
        self.btn_tr_iptal.setStyleSheet("background-color: #F38BA8; color: #11111B; border-radius: 8px; padding: 8px 14px; font-weight: bold;")
        self.btn_tr_iptal.clicked.connect(self.secili_transferi_iptal_et)
        tr_top.addWidget(self.btn_tr_iptal)

        # Özet kartları
        tr_ozet = QHBoxLayout()
        self.lbl_tr_cikis_top = QLabel("→ Toplam Çıkış\n\n0,00 kg")
        self.lbl_tr_cikis_top.setStyleSheet("background-color: #2D1A0A; color: #FAB387; border-radius: 12px; padding: 14px; font-size: 13px; font-weight: bold;")
        self.lbl_tr_cikis_top.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_tr_giris_top = QLabel("← Toplam Giriş\n\n0,00 kg")
        self.lbl_tr_giris_top.setStyleSheet("background-color: #0A2D15; color: #A6E3A1; border-radius: 12px; padding: 14px; font-size: 13px; font-weight: bold;")
        self.lbl_tr_giris_top.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_tr_net = QLabel("Net Bakiye\n\n0,00 kg")
        self.lbl_tr_net.setStyleSheet("background-color: #313244; color: #CDD6F4; border-radius: 12px; padding: 14px; font-size: 13px; font-weight: bold;")
        self.lbl_tr_net.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tr_ozet.addWidget(self.lbl_tr_cikis_top)
        tr_ozet.addWidget(self.lbl_tr_giris_top)
        tr_ozet.addWidget(self.lbl_tr_net)

        transfer_layout.addLayout(tr_top)
        transfer_layout.addLayout(tr_ozet)

        # Transfer tablosu
        self.table_transfer = GelistirilmisTablo(self)
        self.table_transfer.setColumnCount(8)
        self.table_transfer.setHorizontalHeaderLabels(["Tarih", "Barkod", "Kategori", "Miktar (kg)", "Yön", "İşletme", "Kalan (kg)", "ID"])
        self.table_transfer.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_transfer.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_transfer.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_transfer.horizontalHeader().setStretchLastSection(False)
        self.table_transfer.setShowGrid(True)
        self.table_transfer.setStyleSheet("QTableWidget { gridline-color: #313244; }")
        self.table_transfer.itemDoubleClicked.connect(self.transfer_duzenle_ac)
        transfer_layout.addWidget(self.table_transfer)

        self._transfer_filtre = "hepsi"

        # ── KISAYOLLAR SEKMESİ ──
        kisayol_layout = QVBoxLayout(self.tab_kisayollar)
        kisayol_layout.setContentsMargins(10, 10, 10, 10)

        lbl_baslik = QLabel("⌨️ Klavye Kısayolları")
        lbl_baslik.setStyleSheet("font-size: 18px; font-weight: bold; color: #FAB387; margin-bottom: 10px;")
        kisayol_layout.addWidget(lbl_baslik)

        self.table_kisayollar = self._analiz_tablosu(["Kısayol", "İşlev", "Açıklama"])
        kisayol_layout.addWidget(self.table_kisayollar)

        kisayol_data = [
            ("Ctrl + A", "Ana Sayfa", "Ana sayfa sekmesine geç"),
            ("Ctrl + E", "Et Kayıtları", "Et kayıtları sekmesine geç"),
            ("Ctrl + T", "Tavuk Kayıtları", "Tavuk kayıtları sekmesine geç"),
            ("Ctrl + Z", "Analiz", "Analiz & Raporlar sekmesine geç"),
            ("Ctrl + R", "Transfer", "Transfer sekmesine geç"),
            ("Ctrl + K", "Kısayollar", "Bu ekrana gel"),
            ("Ctrl + N", "Yeni Kayıt", "Aktif sekmeye göre yeni kayıt ekle"),
            ("Ctrl + Alt + E", "Et Ekle", "Et sekmesine geç ve yeni et kaydı ekle"),
            ("Ctrl + Alt + T", "Tavuk Ekle", "Tavuk sekmesine geç ve yeni tavuk kaydı ekle"),
            ("Ctrl + D", "Sil", "Seçili kaydı sil"),
            ("F5", "Yenile", "Verileri sunucudan yenile"),
            ("Ctrl + B", "Excel'den Al", "Excel dosyasından veri içe aktar"),
            ("Ctrl + X", "Excel'e Aktar", "Tüm verileri Excel'e aktar"),
            ("Ctrl + Scroll ↑↓", "Punto Boyutu", "Tablo yazı boyutunu büyüt/küçült"),
            ("Çift Tıkla", "Düzenle", "Hücreyi düzenle (tarih/kg/yön için özel giriş)"),
            ("Sağ Tıkla", "Menü", "Düzenle / Kopyala / Transfer / Sil menüsü"),
        ]

        for kisayol, isim, aciklama in kisayol_data:
            r = self.table_kisayollar.rowCount()
            self.table_kisayollar.insertRow(r)
            k_item = QTableWidgetItem(kisayol)
            k_item.setForeground(QColor("#FAB387"))
            k_item.setFont(QFont("Courier New", 12, QFont.Weight.Bold))
            self.table_kisayollar.setItem(r, 0, k_item)
            self.table_kisayollar.setItem(r, 1, QTableWidgetItem(isim))
            self.table_kisayollar.setItem(r, 2, QTableWidgetItem(aciklama))
            for col in range(3):
                item = self.table_kisayollar.item(r, col)
                if item:
                    item.setBackground(QColor("#24273A") if r % 2 == 0 else QColor("#1E1E2E"))

        alt_bar = QHBoxLayout()
        self.lbl_secim_bilgi = QLabel("📊 Seçili Hücre: 0 | Toplam Değer: 0,00")
        self.lbl_secim_bilgi.setObjectName("lbl_secim")
        alt_bar.addWidget(self.lbl_secim_bilgi)
        alt_bar.addStretch()
        self.lbl_durum = QLabel("")
        self.lbl_durum.setStyleSheet("color: #A6E3A1; font-weight: bold; font-size: 15px; padding: 10px;")
        alt_bar.addWidget(self.lbl_durum)
        size_grip = QSizeGrip(self)
        size_grip.setStyleSheet("width: 15px; height: 15px; background-color: transparent;")
        alt_bar.addWidget(size_grip, 0, Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignRight)
        layout.addLayout(alt_bar)

        self.tabs.currentChanged.connect(lambda: self.lbl_secim_bilgi.setText("📊 Seçili Hücre: 0 | Toplam Değer: 0,00"))
        self.shortcut_kaydet = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_kaydet.activated.connect(lambda: self.lbl_durum.setText("☁️ Bulut ile senkronize!") or QTimer.singleShot(3000, lambda: self.lbl_durum.setText("")))

        # ── KISAYOL TUŞLARI ──
        kisayollar = [
            ("Ctrl+A", lambda: self.tabs.setCurrentIndex(0)),           # Ana Sayfa
            ("Ctrl+E", lambda: self.tabs.setCurrentIndex(1)),           # Et Kayıtları
            ("Ctrl+T", lambda: self.tabs.setCurrentIndex(2)),           # Tavuk Kayıtları
            ("Ctrl+Z", lambda: self.tabs.setCurrentIndex(3)),           # Analiz
            ("Ctrl+R", lambda: self.tabs.setCurrentIndex(4)),           # Transfer
            ("Ctrl+K", lambda: self.tabs.setCurrentIndex(5)),           # Kısayollar
            ("Ctrl+N", self.yeni_urun_penceresi_ac),                    # Yeni kayıt
            ("Ctrl+Alt+E", lambda: (self.tabs.setCurrentIndex(1), self.yeni_urun_penceresi_ac())),  # Et ekle
            ("Ctrl+Alt+T", lambda: (self.tabs.setCurrentIndex(2), self.yeni_urun_penceresi_ac())),  # Tavuk ekle
            ("Ctrl+D", self.urun_sil),                                  # Sil
            ("F5", self.verileri_yukle),                                # Yenile
            ("Ctrl+B", self.excelden_iceri_aktar),                      # Excel'den al
            ("Ctrl+X", self.excel_disa_aktar),                         # Excel'e aktar
        ]
        for kisayol, fonksiyon in kisayollar:
            sc = QShortcut(QKeySequence(kisayol), self)
            sc.activated.connect(fonksiyon)

    def _analiz_tablosu(self, basliklar):
        t = GelistirilmisTablo(self)
        t.setColumnCount(len(basliklar))
        t.setHorizontalHeaderLabels(basliklar)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        return t

    def tablo_olustur(self):
        table = GelistirilmisTablo(self)
        table.setColumnCount(13)
        self.basliklar = ["Barkod", "Geliş\nT.", "Kullanım\nT.", "Küvet T.\n(Tekrar)", "Küvet\n(DÜŞ)", "Takoz T.\n(Tekrar)", "Takoz\n(DÜŞ)", "İlk\nMiktar", "Kalan\nMiktar", "Tüketim\n(%)", "Zayi\n(kg)", "Zayi\nT.", "ID"]
        table.setHorizontalHeaderLabels(self.basliklar)
        table.setFont(QFont("Arial", 12))
        table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.EditKeyPressed)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        header_h = table.horizontalHeader()
        header_h.setStretchLastSection(False)
        header_h.setMinimumSectionSize(40)
        header_h.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header_h.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        header_h.setFixedHeight(48)  # İki satır sığacak yükseklik
        table.setShowGrid(True)
        table.setStyleSheet("""
            QTableWidget { gridline-color: #45475A; }
            QHeaderView::section:vertical {
                background-color: #1E1E2E;
                color: #A6ADC8;
                font-size: 11px;
                font-weight: bold;
                min-width: 60px;
                max-width: 60px;
                padding: 0px;
            }
        """)
        table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        table.verticalHeader().setDefaultSectionSize(28)
        # Arial 12pt'de 3 basamaklı sayı ~45px
        table.verticalHeader().setMinimumWidth(60)
        table.verticalHeader().setMaximumWidth(60)
        table.verticalHeader().setDefaultSectionSize(30)
        header = table.horizontalHeader()
        header.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        header.customContextMenuRequested.connect(lambda pos, t=table: self.baslik_menusu_ac(pos, t))
        # Sıralama sonrası renklendirme verileri_yukle sonunda yapılıyor
        # Tarih sütunları için DateDelegate
        _date_delegate = DateDelegate(table)
        for col in [1, 2, 3, 5, 11, 13]:
            table.setItemDelegateForColumn(col, _date_delegate)
        # kg sütunları için SpinBoxDelegate
        _spin_delegate = SpinBoxDelegate(table)
        for col in [4, 6, 7, 8, 10, 12]:  # Küvet DÜŞ, Takoz DÜŞ, İlk, Kalan, Zayi, Transfer kg
            table.setItemDelegateForColumn(col, _spin_delegate)
        # (Transfer sütunları artık bu tabloda yok)
        table.itemChanged.connect(self.hucre_degisti)
        table.itemSelectionChanged.connect(self.secim_hesapla)
        return table

    def secim_hesapla(self):
        tablo = self.aktif_tabloyu_al()
        if not tablo: return
        secili_hucreler = tablo.selectedItems()
        toplam, sayisal_adet = 0.0, 0
        for item in secili_hucreler:
            try:
                toplam += float(item.text().replace("kg", "").replace("%", "").replace("(KRİTİK)", "").strip().replace(',', '.'))
                sayisal_adet += 1
            except ValueError:
                pass
        adet = len(secili_hucreler)
        self.lbl_secim_bilgi.setText(f"📊 Seçili Hücre: {adet} | Toplam Değer: {toplam:.3f}".replace('.', ',') if sayisal_adet > 0 else f"📊 Seçili Hücre: {adet} | Toplam Değer: 0,00")

    def basliklari_gorselle(self):
        for tablo in [self.table_et, self.table_tavuk]:
            for i, baslik in enumerate(self.basliklar):
                item = tablo.horizontalHeaderItem(i)
                if i in self.aktif_filtreler:
                    item.setText(baslik + " 🔽")
                    item.setForeground(QColor("#A6E3A1"))
                else:
                    item.setText(baslik)
                    item.setForeground(QColor("#A6ADC8"))

    def baslik_menusu_ac(self, pos, tablo):
        sutun = tablo.horizontalHeader().logicalIndexAt(pos)
        benzersiz_degerler = {tablo.item(row, sutun).text() for row in range(tablo.rowCount()) if tablo.item(row, sutun)}
        dialog = SutunFiltreDialog(self.basliklar[sutun], benzersiz_degerler, self.aktif_filtreler.get(sutun), self)
        dialog.setStyleSheet(self.styleSheet())
        if dialog.exec() == QDialog.DialogCode.Accepted:
            secilenler = dialog.secilenleri_al()
            if len(secilenler) == len(benzersiz_degerler) or len(secilenler) == 0:
                self.aktif_filtreler.pop(sutun, None)
            else:
                self.aktif_filtreler[sutun] = secilenler
            self.basliklari_gorselle()
            self.tablolari_filtrele()

    def tablolari_filtrele(self, metin=None):
        if metin is None: metin = self.arama_kutusu.text()
        metin = metin.lower().strip()
        for tablo in [self.table_et, self.table_tavuk]:
            for row in range(tablo.rowCount()):
                gizle = False
                for col, izin_verilenler in self.aktif_filtreler.items():
                    item = tablo.item(row, col)
                    if item and item.text() not in izin_verilenler:
                        gizle = True; break
                if not gizle and metin:
                    gizle = not any(
                        metin in (tablo.item(row, col).text().lower() if tablo.item(row, col) else "")
                        for col in range(tablo.columnCount())
                    )
                tablo.setRowHidden(row, gizle)

    def aktif_tabloyu_al(self):
        idx = self.tabs.currentIndex()
        if idx == 1: return self.table_et
        elif idx == 2: return self.table_tavuk
        return None

    def tarih_gercek_degeri(self, t_str):
        if not t_str or t_str == "-": return "-"
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try: return datetime.strptime(t_str, fmt)
            except: pass
        return "-"

    # ── VERİ YÜKLEME (API'den) ──
    def _baglanti_hatasi(self, hata):
        self.lbl_durum.setText(f"❌ {hata}")
        # Yerel yedekten yükle
        yedek = self._yerel_yedekten_yukle()
        if yedek:
            cevap = QMessageBox.question(self, "Bağlantı Hatası",
                f"Sunucuya bağlanılamadı.\n"
                f"Son yedek: {yedek.get('tarih', '?')} — {yedek.get('kayit_sayisi', 0)} kayıt\n\n"
                "Yerel yedekten yüklensin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if cevap == QMessageBox.StandardButton.Yes:
                self._verileri_isle(yedek["veriler"])
                self.lbl_durum.setText("📂 Yerel yedekten yüklendi")
                QTimer.singleShot(5000, lambda: self.lbl_durum.setText(""))
        else:
            QTimer.singleShot(5000, lambda: self.lbl_durum.setText(""))

    def verileri_yukle(self):
        self.lbl_durum.setText("☁️ Yükleniyor...")
        # Önceki thread hâlâ çalışıyorsa bekle
        if hasattr(self, 'yukleyici') and self.yukleyici and self.yukleyici.isRunning():
            return
        self.yukleyici = VeriYukleyici(self.api)
        self.yukleyici.veri_hazir.connect(self._verileri_isle)
        self.yukleyici.hata_olustu.connect(self._baglanti_hatasi)
        self.yukleyici.finished.connect(lambda: None)  # referans tut
        self.yukleyici.start()

    def _verileri_isle(self, veriler_raw):
        # API'den gelen dict listesini tuple listesine çevir (mevcut kod uyumlu)
        veriler = [
            (u["id"], u["barkod"], u["kategori"], u["gelis_tarihi"],
             u.get("kullanim_tarihi", "-"),
             u.get("kuvet_kullanim_tarihi", "-"), u.get("kuvet_miktar", 0.0),
             u.get("takoz_kullanim_tarihi", "-"), u.get("takoz_miktar", 0.0),
             u["ilk_miktar"], u["kalan_miktar"], u.get("zayi_miktar", 0.0),
             u.get("zayi_tarihi", "-"), u.get("transfer_miktar", 0.0),
             u.get("transfer_tarihi", "-"), u.get("transfer_yon", "-"),
             u.get("transfer_isletme", "-"))
            for u in veriler_raw
        ]

        self.table_et.blockSignals(True)
        self.table_tavuk.blockSignals(True)
        self.table_et.setSortingEnabled(False)
        self.table_tavuk.setSortingEnabled(False)
        self.table_et.setRowCount(0)
        self.table_tavuk.setRowCount(0)

        et_giren, et_kalan, tavuk_giren, tavuk_kalan = 0.0, 0.0, 0.0, 0.0
        son_7_et, son_7_tavuk, son_30_et, son_30_tavuk = 0.0, 0.0, 0.0, 0.0
        def _bos_veri():
            return {"Et": 0.0, "Tavuk": 0.0, "Zayi": 0.0,
                    "paket": {"4": 0, "6": 0, "8": 0, "10": 0, "12": 0,
                               "15": 0, "20": 0, "30": 0, "40": 0, "50": 0, "diger": 0}}
        self.gunluk_veri = defaultdict(_bos_veri)
        self.haftalik_veri = defaultdict(_bos_veri)
        self.aylik_veri = defaultdict(_bos_veri)

        # Paket raporu: {(donem, paket_kg): {Et_adet, Et_kg, Tavuk_adet, Tavuk_kg}}
        _pk_bos = lambda: {"Et_adet": 0, "Et_kg": 0.0, "Tavuk_adet": 0, "Tavuk_kg": 0.0}
        self.paket_gun_veri = defaultdict(_pk_bos)   # key: (gun_str, pk)
        self.paket_ay_veri  = defaultdict(_pk_bos)   # key: (ay_str,  pk)
        # Genel: {pk: {giren_adet, giren_kg, kullanilan_adet, kullanilan_kg, kalan_adet, kalan_kg}}
        _pk_genel_bos = lambda: {"giren_adet": 0, "giren_kg": 0.0,
                                  "kullanilan_adet": 0, "kullanilan_kg": 0.0,
                                  "kalan_adet": 0, "kalan_kg": 0.0}
        self.paket_genel_veri = defaultdict(_pk_genel_bos)
        self.gelen_koli_veri = defaultdict(lambda: {
            "4_adet": 0, "4_kg": 0.0, "6_adet": 0, "6_kg": 0.0,
            "8_adet": 0, "8_kg": 0.0, "10_adet": 0, "10_kg": 0.0,
            "12_adet": 0, "12_kg": 0.0, "15_adet": 0, "15_kg": 0.0,
            "20_adet": 0, "20_kg": 0.0, "30_adet": 0, "30_kg": 0.0,
            "40_adet": 0, "40_kg": 0.0, "50_adet": 0, "50_kg": 0.0,
            "diger_adet": 0, "diger_kg": 0.0})
        bugun = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        for row_data in veriler:
            _id, barkod, kategori, gelis, kullanim, kuvet_tar, kuvet_mik, takoz_tar, takoz_mik, ilk_mik, kalan_mik, zayi_mik, zayi_tar, transfer_mik, transfer_tar, transfer_yon, transfer_isletme = row_data

            if gelis != "-":
                try:
                    g_obj = datetime.strptime(gelis, "%d.%m.%Y")
                    g_str = g_obj.strftime("%d.%m.%Y")
                    _gk = self.gelen_koli_veri[(g_str, kategori)]
                    if 3 <= ilk_mik < 5:      _gk["4_adet"]  += 1; _gk["4_kg"]  += ilk_mik; _gpk = "4"
                    elif 5 <= ilk_mik < 7:    _gk["6_adet"]  += 1; _gk["6_kg"]  += ilk_mik; _gpk = "6"
                    elif 7 <= ilk_mik < 9:    _gk["8_adet"]  += 1; _gk["8_kg"]  += ilk_mik; _gpk = "8"
                    elif 9 <= ilk_mik <= 11:  _gk["10_adet"] += 1; _gk["10_kg"] += ilk_mik; _gpk = "10"
                    elif 11 < ilk_mik <= 13:  _gk["12_adet"] += 1; _gk["12_kg"] += ilk_mik; _gpk = "12"
                    elif 14 <= ilk_mik <= 17: _gk["15_adet"] += 1; _gk["15_kg"] += ilk_mik; _gpk = "15"
                    elif 19 <= ilk_mik <= 21: _gk["20_adet"] += 1; _gk["20_kg"] += ilk_mik; _gpk = "20"
                    elif 28 <= ilk_mik <= 32: _gk["30_adet"] += 1; _gk["30_kg"] += ilk_mik; _gpk = "30"
                    elif 38 <= ilk_mik <= 42: _gk["40_adet"] += 1; _gk["40_kg"] += ilk_mik; _gpk = "40"
                    elif 48 <= ilk_mik <= 52: _gk["50_adet"] += 1; _gk["50_kg"] += ilk_mik; _gpk = "50"
                    else:                      _gk["diger_adet"] += 1; _gk["diger_kg"] += ilk_mik; _gpk = "diger"
                    # Genel paket: giren ve kalan
                    self.paket_genel_veri[_gpk]["giren_adet"] += 1
                    self.paket_genel_veri[_gpk]["giren_kg"]   += ilk_mik
                    self.paket_genel_veri[_gpk]["kalan_adet"] += 1 if kalan_mik > 0 else 0
                    self.paket_genel_veri[_gpk]["kalan_kg"]   += kalan_mik
                except ValueError: pass

            # İlk kullanım tüketimi (küvet ve takoz hariç)
            ilk_kullanim_mik = max(0.0, ilk_mik - kalan_mik - kuvet_mik - takoz_mik - zayi_mik - transfer_mik)
            if ilk_kullanim_mik > 0 and kullanim != "-":
                try:
                    t_obj = None
                    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                        try: t_obj = datetime.strptime(kullanim, fmt); break
                        except: pass
                    if not t_obj: raise ValueError()
                    gun_str = t_obj.strftime('%d.%m.%Y')
                    hafta_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%W')}. Hafta"
                    ay_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%m')} ({self.aylar.get(t_obj.strftime('%m'), '')})"
                    self.gunluk_veri[gun_str][kategori] += ilk_kullanim_mik
                    self.haftalik_veri[hafta_str][kategori] += ilk_kullanim_mik
                    self.aylik_veri[ay_str][kategori] += ilk_kullanim_mik
                    # Paket kg kategorisi tespiti
                    if 3 <= ilk_mik < 5:       _pk = "4"
                    elif 5 <= ilk_mik < 7:     _pk = "6"
                    elif 7 <= ilk_mik < 9:     _pk = "8"
                    elif 9 <= ilk_mik <= 11:   _pk = "10"
                    elif 11 < ilk_mik <= 13:   _pk = "12"
                    elif 14 <= ilk_mik <= 17:  _pk = "15"
                    elif 19 <= ilk_mik <= 21:  _pk = "20"
                    elif 28 <= ilk_mik <= 32:  _pk = "30"
                    elif 38 <= ilk_mik <= 42:  _pk = "40"
                    elif 48 <= ilk_mik <= 52:  _pk = "50"
                    else:                       _pk = "diger"
                    self.gunluk_veri[gun_str]["paket"][_pk] += 1
                    self.haftalik_veri[hafta_str]["paket"][_pk] += 1
                    self.aylik_veri[ay_str]["paket"][_pk] += 1
                    # Paket raporu: günlük ve aylık kullanım
                    _kat = kategori  # "Et" veya "Tavuk"
                    self.paket_gun_veri[(gun_str, _pk)][f"{_kat}_adet"] += 1
                    self.paket_gun_veri[(gun_str, _pk)][f"{_kat}_kg"]   += ilk_kullanim_mik
                    self.paket_ay_veri[(ay_str, _pk)][f"{_kat}_adet"]   += 1
                    self.paket_ay_veri[(ay_str, _pk)][f"{_kat}_kg"]     += ilk_kullanim_mik
                    # Genel: kullanılan
                    self.paket_genel_veri[_pk]["kullanilan_adet"] += 1
                    self.paket_genel_veri[_pk]["kullanilan_kg"]   += ilk_kullanim_mik
                    fark_gun = (bugun - t_obj).days
                    if 0 <= fark_gun <= 7:
                        if kategori == "Et": son_7_et += ilk_kullanim_mik
                        else: son_7_tavuk += ilk_kullanim_mik
                    if 0 <= fark_gun <= 30:
                        if kategori == "Et": son_30_et += ilk_kullanim_mik
                        else: son_30_tavuk += ilk_kullanim_mik
                except: pass
            # Küvet tüketimi analizi
            if kuvet_mik > 0 and kuvet_tar != "-":
                try:
                    t_obj = None
                    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                        try: t_obj = datetime.strptime(kuvet_tar, fmt); break
                        except: pass
                    if not t_obj: raise ValueError()
                    gun_str = t_obj.strftime('%d.%m.%Y')
                    hafta_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%W')}. Hafta"
                    ay_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%m')} ({self.aylar.get(t_obj.strftime('%m'), '')})"
                    self.gunluk_veri[gun_str][kategori] += kuvet_mik
                    self.haftalik_veri[hafta_str][kategori] += kuvet_mik
                    self.aylik_veri[ay_str][kategori] += kuvet_mik
                    fark_gun = (bugun - t_obj).days
                    if 0 <= fark_gun <= 7:
                        if kategori == "Et": son_7_et += kuvet_mik
                        else: son_7_tavuk += kuvet_mik
                    if 0 <= fark_gun <= 30:
                        if kategori == "Et": son_30_et += kuvet_mik
                        else: son_30_tavuk += kuvet_mik
                except: pass
            # Takoz tüketimi analizi
            if takoz_mik > 0 and takoz_tar != "-":
                try:
                    t_obj = None
                    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                        try: t_obj = datetime.strptime(takoz_tar, fmt); break
                        except: pass
                    if not t_obj: raise ValueError()
                    gun_str = t_obj.strftime('%d.%m.%Y')
                    hafta_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%W')}. Hafta"
                    ay_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%m')} ({self.aylar.get(t_obj.strftime('%m'), '')})"
                    self.gunluk_veri[gun_str][kategori] += takoz_mik
                    self.haftalik_veri[hafta_str][kategori] += takoz_mik
                    self.aylik_veri[ay_str][kategori] += takoz_mik
                    fark_gun = (bugun - t_obj).days
                    if 0 <= fark_gun <= 7:
                        if kategori == "Et": son_7_et += takoz_mik
                        else: son_7_tavuk += takoz_mik
                    if 0 <= fark_gun <= 30:
                        if kategori == "Et": son_30_et += takoz_mik
                        else: son_30_tavuk += takoz_mik
                except: pass
            # Zayi analizi (tarihe göre, tüketime dahil edilmez — sadece zayi tablosunda)
            if zayi_mik > 0 and zayi_tar != "-":
                try:
                    t_obj = datetime.strptime(zayi_tar, "%d.%m.%Y")
                    gun_str = t_obj.strftime('%d.%m.%Y')
                    hafta_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%W')}. Hafta"
                    ay_str = f"{t_obj.strftime('%Y')} - {t_obj.strftime('%m')} ({self.aylar.get(t_obj.strftime('%m'), '')})"
                    self.gunluk_veri[gun_str]["Zayi"] += zayi_mik
                    self.haftalik_veri[hafta_str]["Zayi"] += zayi_mik
                    self.aylik_veri[ay_str]["Zayi"] += zayi_mik
                except: pass

            if kategori == "Et":
                hedef_tablo = self.table_et
                et_giren += ilk_mik; et_kalan += kalan_mik
            else:
                hedef_tablo = self.table_tavuk
                tavuk_giren += ilk_mik; tavuk_kalan += kalan_mik

            # Transfer olan ürünler normal tabloya eklenmez
            if transfer_mik > 0 and transfer_yon in ("Çıkış", "Giriş"):
                continue

            row_idx = hedef_tablo.rowCount()
            hedef_tablo.insertRow(row_idx)
            tuketim_yuzdesi = (max(0, ilk_mik - kalan_mik - zayi_mik - transfer_mik) / ilk_mik) * 100 if ilk_mik > 0 else 0

            hedef_tablo.setItem(row_idx, 0, SiralanabilirItem(barkod, barkod))
            hedef_tablo.setItem(row_idx, 1, SiralanabilirItem(gelis, self.tarih_gercek_degeri(gelis)))
            hedef_tablo.setItem(row_idx, 2, SiralanabilirItem(kullanim, self.tarih_gercek_degeri(kullanim)))
            hedef_tablo.setItem(row_idx, 3, SiralanabilirItem(kuvet_tar, self.tarih_gercek_degeri(kuvet_tar)))
            hedef_tablo.setItem(row_idx, 4, SiralanabilirItem(f"{kuvet_mik:.3f} kg".replace('.', ','), kuvet_mik))
            hedef_tablo.setItem(row_idx, 5, SiralanabilirItem(takoz_tar, self.tarih_gercek_degeri(takoz_tar)))
            hedef_tablo.setItem(row_idx, 6, SiralanabilirItem(f"{takoz_mik:.3f} kg".replace('.', ','), takoz_mik))
            hedef_tablo.setItem(row_idx, 7, SiralanabilirItem(f"{ilk_mik:.3f} kg".replace('.', ','), ilk_mik))
            kalan_item = SiralanabilirItem(f"{kalan_mik:.3f} kg".replace('.', ','), kalan_mik)
            if kalan_mik > 0 and kalan_mik <= (ilk_mik * 0.20): kalan_item.setForeground(QColor("#F38BA8"))
            elif kalan_mik == 0: kalan_item.setForeground(QColor("#A6ADC8"))
            hedef_tablo.setItem(row_idx, 8, kalan_item)
            tuk_item = SiralanabilirItem(f"% {tuketim_yuzdesi:.0f}", tuketim_yuzdesi)
            if tuketim_yuzdesi >= 90: tuk_item.setForeground(QColor("#F38BA8"))   # kırmızı
            elif tuketim_yuzdesi >= 70: tuk_item.setForeground(QColor("#FAB387"))  # turuncu
            elif tuketim_yuzdesi >= 40: tuk_item.setForeground(QColor("#F9E2AF"))  # sarı
            else: tuk_item.setForeground(QColor("#A6E3A1"))                        # yeşil
            hedef_tablo.setItem(row_idx, 9, tuk_item)
            zayi_item = SiralanabilirItem(f"{zayi_mik:.3f} kg".replace('.', ','), zayi_mik)
            if zayi_mik > 0: zayi_item.setForeground(QColor("#F9E2AF"))
            hedef_tablo.setItem(row_idx, 10, zayi_item)
            hedef_tablo.setItem(row_idx, 11, SiralanabilirItem(zayi_tar, self.tarih_gercek_degeri(zayi_tar)))
            # Transfer
            id_item = SiralanabilirItem(str(_id), _id)
            hedef_tablo.setItem(row_idx, 12, id_item)

            for col in range(13):
                item = hedef_tablo.item(row_idx, col)
                if item:
                    if transfer_mik > 0 and transfer_yon == "Çıkış":
                        item.setBackground(QColor("#3D1F05"))
                        # Transfer olan satır: sadece 12-15 düzenlenebilir
                        if col in [12, 13, 14, 15]:
                            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                        else:
                            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    elif transfer_mik > 0 and transfer_yon == "Giriş":
                        item.setBackground(QColor("#063D15"))
                        if col in [12, 13, 14, 15]:
                            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                        else:
                            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    else:
                        item.setBackground(QColor("#24273A") if row_idx % 2 == 0 else QColor("#1E1E2E"))
                        # Normal satır: standart düzenlenebilir sütunlar
                        if col in [0, 1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 13, 14, 15]:
                            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                        else:
                            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

        self.lbl_et_toplam.setText(f"🥩 ET STOK DURUMU\n\nGiren: {et_giren:.3f} kg\nKalan: {et_kalan:.3f} kg".replace('.', ','))
        self.lbl_tavuk_toplam.setText(f"🍗 TAVUK STOK DURUMU\n\nGiren: {tavuk_giren:.3f} kg\nKalan: {tavuk_kalan:.3f} kg".replace('.', ','))
        self.lbl_ort_7.setText(f"⏳ Son 7 Günlük Tüketim Ortalaması\n\nEt: {(son_7_et/7):.3f} kg/gün\nTavuk: {(son_7_tavuk/7):.3f} kg/gün".replace('.', ','))
        self.lbl_ort_30.setText(f"📅 Son 30 Günlük Tüketim Ortalaması\n\nEt: {(son_30_et/30):.3f} kg/gün\nTavuk: {(son_30_tavuk/30):.3f} kg/gün".replace('.', ','))

        self.table_et.setSortingEnabled(True)
        self.table_tavuk.setSortingEnabled(True)

        if self.ilk_acilis:
            et_col = int(self.ayarlar.value("et_sort_col", 0))
            et_order = int(self.ayarlar.value("et_sort_order", Qt.SortOrder.DescendingOrder.value))
            tavuk_col = int(self.ayarlar.value("tavuk_sort_col", 0))
            tavuk_order = int(self.ayarlar.value("tavuk_sort_order", Qt.SortOrder.DescendingOrder.value))
            self.table_et.sortItems(et_col, Qt.SortOrder(et_order))
            self.table_tavuk.sortItems(tavuk_col, Qt.SortOrder(tavuk_order))
            self.ilk_acilis = False
        else:
            et_col = self.table_et.horizontalHeader().sortIndicatorSection()
            if et_col >= 0: self.table_et.sortItems(et_col, self.table_et.horizontalHeader().sortIndicatorOrder())
            tavuk_col = self.table_tavuk.horizontalHeader().sortIndicatorSection()
            if tavuk_col >= 0: self.table_tavuk.sortItems(tavuk_col, self.table_tavuk.horizontalHeader().sortIndicatorOrder())

        self.table_et.blockSignals(False)
        self.table_tavuk.blockSignals(False)
        self._sutun_genisliklerini_ayarla()
        self.basliklari_gorselle()
        self.satirlari_renklendir(self.table_et)
        self.satirlari_renklendir(self.table_tavuk)
        self.tablolari_filtrele()
        self.analiz_tablolarini_doldur()
        self._stok_ozet_doldur(veriler_raw)
        self.transfer_tablosunu_doldur()
        self.lbl_durum.setText("✅ Senkronize edildi")
        QTimer.singleShot(3000, lambda: self.lbl_durum.setText(""))
        # Yerel yedek kaydet
        self._yerel_yedek_kaydet(veriler_raw)

    def transfer_tablosunu_doldur(self):
        """Transfer sekmesi tablosunu ve özet kartlarını güncelle"""
        self.table_transfer.setSortingEnabled(False)
        self.table_transfer.setRowCount(0)
        filtre = self._transfer_filtre
        toplam_cikis = toplam_giris = 0.0

        # Tüm ürünlerden transfer olanları topla
        try:
            veriler_raw = self.api.urunleri_getir()
        except Exception:
            return

        for u in veriler_raw:
            tr_mik = u.get("transfer_miktar", 0.0) or 0.0
            tr_yon = u.get("transfer_yon", "-") or "-"
            if tr_mik <= 0 or tr_yon == "-":
                continue
            if filtre != "hepsi" and tr_yon != filtre:
                continue

            if tr_yon == "Çıkış": toplam_cikis += tr_mik
            else: toplam_giris += tr_mik

            tr_tar = u.get("transfer_tarihi", "-") or "-"
            tr_isl = u.get("transfer_isletme", "-") or "-"
            barkod = u.get("barkod", "")
            kat = u.get("kategori", "")
            kalan = u.get("kalan_miktar", 0.0) or 0.0
            uid = u.get("id", 0)

            r = self.table_transfer.rowCount()
            self.table_transfer.insertRow(r)

            renk = QColor("#FAB387") if tr_yon == "Çıkış" else QColor("#A6E3A1")
            bg_renk = QColor("#2D1A0A") if tr_yon == "Çıkış" else QColor("#0A2D15")

            items = [
                QTableWidgetItem(tr_tar),
                QTableWidgetItem(barkod),
                QTableWidgetItem(kat),
                QTableWidgetItem(f"{tr_mik:.3f} kg".replace('.', ',')),
                QTableWidgetItem(tr_yon),
                QTableWidgetItem(tr_isl),
                QTableWidgetItem(f"{kalan:.3f} kg".replace('.', ',')),
                QTableWidgetItem(str(uid)),
            ]
            for col, item in enumerate(items):
                item.setBackground(bg_renk)
                if col in [3, 4]: item.setForeground(renk)
                self.table_transfer.setItem(r, col, item)

        self.table_transfer.setSortingEnabled(True)
        self.table_transfer.resizeColumnsToContents()

        # Özet kartları
        net = toplam_giris - toplam_cikis
        self.lbl_tr_cikis_top.setText(f"→ Toplam Çıkış\n\n{toplam_cikis:.3f} kg".replace('.', ','))
        self.lbl_tr_giris_top.setText(f"← Toplam Giriş\n\n{toplam_giris:.3f} kg".replace('.', ','))
        net_renk = "#A6E3A1" if net >= 0 else "#FAB387"
        self.lbl_tr_net.setStyleSheet(f"background-color: #313244; color: {net_renk}; border-radius: 12px; padding: 14px; font-size: 13px; font-weight: bold;")
        isaret = "+" if net >= 0 else ""
        self.lbl_tr_net.setText(f"Net Bakiye\n\n{isaret}{net:.3f} kg".replace('.', ','))

    def _duzenle_dialog_ac(self, urun_id):
        """Sağ tık Düzenle → tüm alanların düzenlenebileceği dialog"""
        # Ürün verisini bul
        try:
            veriler = self.api.urunleri_getir()
            u = next((x for x in veriler if x["id"] == urun_id), None)
            if not u:
                QMessageBox.warning(self, "Hata", "Ürün bulunamadı!")
                return
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Düzenle — {u.get('barkod','')}")
        dialog.resize(420, 520)
        dialog.setStyleSheet(self.styleSheet())
        layout = QVBoxLayout(dialog)

        ortak = "background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 10px; padding: 6px;"

        form = QFormLayout()
        form.setSpacing(10)

        def tarih_widget(val):
            w = QDateEdit()
            w.setCalendarPopup(True)
            w.setStyleSheet(ortak)
            if val and val != "-":
                try:
                    d, m, y = val.split(".")
                    w.setDate(QDate(int(y), int(m), int(d)))
                except:
                    w.setDate(QDate.currentDate())
            else:
                w.setDate(QDate.currentDate())
            return w

        def kg_widget(val):
            w = QDoubleSpinBox()
            w.setRange(0, 99999)
            w.setDecimals(3)
            w.setSuffix(" kg")
            w.setStyleSheet(ortak)
            try: w.setValue(float(val or 0))
            except: w.setValue(0)
            return w

        f_kul = tarih_widget(u.get("kullanim_tarihi", "-"))
        f_kuvet_tar = tarih_widget(u.get("kuvet_kullanim_tarihi", "-"))
        f_kuvet_kg = kg_widget(u.get("kuvet_miktar", 0))
        f_takoz_tar = tarih_widget(u.get("takoz_kullanim_tarihi", "-"))
        f_takoz_kg = kg_widget(u.get("takoz_miktar", 0))
        f_zayi_kg = kg_widget(u.get("zayi_miktar", 0))
        f_kalan = kg_widget(u.get("kalan_miktar", 0))

        form.addRow("Kullanım Tarihi:", f_kul)
        form.addRow("Küvet Tarihi:", f_kuvet_tar)
        form.addRow("Küvet (kg):", f_kuvet_kg)
        form.addRow("Takoz Tarihi:", f_takoz_tar)
        form.addRow("Takoz (kg):", f_takoz_kg)
        form.addRow("Zayi (kg):", f_zayi_kg)
        form.addRow("Kalan (kg):", f_kalan)

        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(dialog.reject)
        btn_kaydet = QPushButton("💾 Kaydet")
        btn_kaydet.setStyleSheet("background-color: #A6E3A1; color: #11111B; border-radius: 10px; padding: 10px; font-weight: bold;")
        btn_kaydet.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_iptal)
        btn_layout.addWidget(btn_kaydet)
        layout.addLayout(btn_layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            def dmy(w): return w.date().toString("dd.MM.yyyy")
            try:
                guncelleme = {
                    "kullanim_tarihi": dmy(f_kul),
                    "kuvet_kullanim_tarihi": dmy(f_kuvet_tar),
                    "kuvet_miktar": f_kuvet_kg.value(),
                    "takoz_kullanim_tarihi": dmy(f_takoz_tar),
                    "takoz_miktar": f_takoz_kg.value(),
                    "zayi_miktar": f_zayi_kg.value(),
                    "kalan_miktar": f_kalan.value(),
                }
                self.api.urun_guncelle(urun_id, guncelleme)
                self.verileri_yukle()
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))

    def _transfer_dialog_ac(self, urun_id, satir, tablo):
        """Sağ tık Transfer Yap diyalogu"""
        from PyQt6.QtWidgets import QDialog, QFormLayout, QComboBox
        dialog = QDialog(self)
        dialog.setWindowTitle("Transfer Yap")
        dialog.resize(340, 220)
        dialog.setStyleSheet(self.styleSheet())
        layout = QVBoxLayout(dialog)
        form = QFormLayout()

        miktar_input = QDoubleSpinBox()
        miktar_input.setRange(0.01, 99999)
        miktar_input.setDecimals(3)
        miktar_input.setSuffix(" kg")
        miktar_input.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 10px; padding: 6px;")

        tarih_input = QDateEdit()
        tarih_input.setCalendarPopup(True)
        tarih_input.setDate(QDate.currentDate())
        tarih_input.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 10px; padding: 6px;")

        yon_input = QComboBox()
        yon_input.addItems(["Çıkış", "Giriş"])
        yon_input.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 10px; padding: 6px;")

        isletme_input = QLineEdit()
        isletme_input.setPlaceholderText("İşletme adı...")
        isletme_input.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 10px; padding: 8px;")

        form.addRow("Miktar:", miktar_input)
        form.addRow("Tarih:", tarih_input)
        form.addRow("Yön:", yon_input)
        form.addRow("İşletme:", isletme_input)
        layout.addLayout(form)

        btn_layout = QHBoxLayout()
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(dialog.reject)
        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.setStyleSheet("background-color: #A6E3A1; color: #11111B; border-radius: 10px; padding: 10px; font-weight: bold;")
        btn_kaydet.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_iptal)
        btn_layout.addWidget(btn_kaydet)
        layout.addLayout(btn_layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            mik = miktar_input.value()
            tar = tarih_input.date().toString("dd.MM.yyyy")
            yon = yon_input.currentText()
            isl = isletme_input.text().strip() or "-"
            try:
                ilk = float((tablo.item(satir,7) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                kuvet = float((tablo.item(satir,4) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                takoz = float((tablo.item(satir,6) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                zayi = float((tablo.item(satir,10) or QTableWidgetItem("0")).text().replace("kg","").replace(",",".").strip())
                yeni_kalan = max(0.0, ilk - kuvet - takoz - zayi - mik)
                self.api.urun_guncelle(urun_id, {
                    "transfer_miktar": mik, "transfer_tarihi": tar,
                    "transfer_yon": yon, "transfer_isletme": isl,
                    "kalan_miktar": yeni_kalan
                })
                self.verileri_yukle()
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))

    def secili_transferi_iptal_et(self):
        """Transfer tablosunda seçili satırın transferini iptal et"""
        secili = self.table_transfer.selectedItems()
        if not secili:
            QMessageBox.warning(self, "Uyarı", "Lütfen iptal etmek istediğiniz transferi seçin.")
            return
        satir = secili[0].row()
        uid_item = self.table_transfer.item(satir, 7)
        if not uid_item: return
        urun_id = int(uid_item.text())
        barkod = self.table_transfer.item(satir, 1).text() if self.table_transfer.item(satir, 1) else ""
        tr_mik = self.table_transfer.item(satir, 3).text() if self.table_transfer.item(satir, 3) else ""
        tr_yon = self.table_transfer.item(satir, 4).text() if self.table_transfer.item(satir, 4) else ""
        tr_isl = self.table_transfer.item(satir, 5).text() if self.table_transfer.item(satir, 5) else ""

        mesaj = f"Barkod: {barkod}\n{tr_yon}: {tr_mik} — {tr_isl}\n\nBu transferi iptal etmek istiyor musunuz?\nÜrün stoka geri dönecek."
        cevap = QMessageBox.question(self, "Transfer İptal", mesaj,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if cevap != QMessageBox.StandardButton.Yes:
            return
        try:
            veriler = self.api.urunleri_getir()
            u = next((x for x in veriler if x["id"] == urun_id), None)
            if not u:
                QMessageBox.warning(self, "Hata", "Ürün bulunamadı!")
                return
            ilk = u.get("ilk_miktar", 0) or 0
            kuvet = u.get("kuvet_miktar", 0) or 0
            takoz = u.get("takoz_miktar", 0) or 0
            zayi = u.get("zayi_miktar", 0) or 0
            yeni_kalan = max(0.0, ilk - kuvet - takoz - zayi)
            self.api.urun_guncelle(urun_id, {
                "transfer_miktar": 0.0,
                "transfer_tarihi": "-",
                "transfer_yon": "-",
                "transfer_isletme": "-",
                "kalan_miktar": yeni_kalan
            })
            self.verileri_yukle()
            QMessageBox.information(self, "Başarılı", f"Transfer iptal edildi. Kalan: {yeni_kalan:.3f} kg")
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))

    def transfer_filtrele(self, filtre):
        self._transfer_filtre = filtre
        # Buton stillerini güncelle
        aktif = "background-color: #FAB387; color: #11111B; border-radius: 8px; padding: 8px 14px; font-weight: bold;"
        pasif = "background-color: #313244; color: #CDD6F4; border-radius: 8px; padding: 8px 14px;"
        self.btn_tr_tumü.setStyleSheet(aktif if filtre == "hepsi" else pasif)
        self.btn_tr_cikis.setStyleSheet(aktif if filtre == "Çıkış" else pasif)
        self.btn_tr_giris.setStyleSheet(aktif if filtre == "Giriş" else pasif)
        self.transfer_tablosunu_doldur()

    def transfer_duzenle_ac(self, item):
        """Transfer satırına çift tıklanınca ürünü modalda aç"""
        row = item.row()
        uid_item = self.table_transfer.item(row, 7)
        if not uid_item: return
        urun_id = int(uid_item.text())
        # İlgili ürünü Et veya Tavuk tablosunda bul ve modalı aç
        for tablo in [self.table_et, self.table_tavuk]:
            for r in range(tablo.rowCount()):
                id_item = tablo.item(r, 12)
                if id_item and int(id_item.text()) == urun_id:
                    msg = (f"ID: {urun_id}\n"
                           f"Barkod: {self.table_transfer.item(row,1).text()}\n"
                           f"Transfer: {self.table_transfer.item(row,3).text()} {self.table_transfer.item(row,4).text()}\n"
                           f"İşletme: {self.table_transfer.item(row,5).text()}")
                    QMessageBox.information(self, "Transfer Kaydı", msg)
                    return

    def satirlari_renklendir(self, tablo):
        """Bej / yeşil zebra - yazı rengi arka plana göre"""
        tablo.blockSignals(True)
        # Arka plan renkleri
        bg_bej   = QColor("#3A3520")   # koyu amber/bej
        bg_yesil = QColor("#142814")   # koyu orman yeşili
        # Yazı renkleri - arka planla belirgin zıt
        fg_bej   = QColor("#7FFFD4")   # aquamarine — bej üstünde belirgin
        fg_yesil = QColor("#FFD700")   # altın sarısı — koyu yeşil üstünde belirgin
        col_count = tablo.columnCount()
        for row in range(tablo.rowCount()):
            bg = bg_bej if row % 2 == 0 else bg_yesil
            fg = fg_bej if row % 2 == 0 else fg_yesil
            for col in range(col_count):
                item = tablo.item(row, col)
                if item:
                    item.setBackground(bg)
                    # Özel renkli sütunları koru (kalan kritik, zayi, tüketim %)
                    mevcut_fg = item.foreground().color()
                    ozel_renkler = [
                        QColor("#F38BA8"), QColor("#A6E3A1"),  # kritik kırmızı, yeşil
                        QColor("#F9E2AF"), QColor("#FAB387"),  # sarı, turuncu
                        QColor("#CBA6F7"), QColor("#89B4FA"),  # mor, mavi
                        QColor("#A6ADC8"),                     # gri barkod
                    ]
                    if not any(mevcut_fg.name() == r.name() for r in ozel_renkler):
                        item.setForeground(fg)
        tablo.blockSignals(False)

    def analiz_tablolarini_doldur(self):
        self.table_gunluk.setRowCount(0)
        self.table_haftalik.setRowCount(0)
        self.table_aylik.setRowCount(0)
        self.table_gelen.setRowCount(0)
        self.table_stok_ozet.setRowCount(0)


        def _paket_str(paket_dict):
            parcalar = []
            for kg in ["4","6","8","10","12","15","20","30","40","50","diger"]:
                adet = paket_dict.get(kg, 0)
                if adet > 0:
                    etiket = "Diğer" if kg == "diger" else f"{kg}kg"
                    parcalar.append(f"{adet}x{etiket}")
            return ", ".join(parcalar) if parcalar else "-"

        for gun_key in sorted(self.gunluk_veri.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"), reverse=True):
            r = self.table_gunluk.rowCount(); self.table_gunluk.insertRow(r)
            et = self.gunluk_veri[gun_key]["Et"]; tavuk = self.gunluk_veri[gun_key]["Tavuk"]; zayi = self.gunluk_veri[gun_key]["Zayi"]
            self.table_gunluk.setItem(r, 0, QTableWidgetItem(gun_key))
            self.table_gunluk.setItem(r, 1, QTableWidgetItem(f"{et:.3f} kg".replace('.', ',')))
            self.table_gunluk.setItem(r, 2, QTableWidgetItem(f"{tavuk:.3f} kg".replace('.', ',')))
            self.table_gunluk.setItem(r, 3, QTableWidgetItem(f"{(et+tavuk):.3f} kg".replace('.', ',')))
            zayi_item = QTableWidgetItem(f"{zayi:.3f} kg".replace('.', ','))
            if zayi > 0: zayi_item.setForeground(QColor("#F9E2AF"))
            self.table_gunluk.setItem(r, 4, zayi_item)
            self.table_gunluk.setItem(r, 5, QTableWidgetItem(_paket_str(self.gunluk_veri[gun_key]["paket"])))

        for h_key in sorted(self.haftalik_veri.keys(), reverse=True):
            r = self.table_haftalik.rowCount(); self.table_haftalik.insertRow(r)
            et = self.haftalik_veri[h_key]["Et"]; tavuk = self.haftalik_veri[h_key]["Tavuk"]; zayi = self.haftalik_veri[h_key]["Zayi"]
            self.table_haftalik.setItem(r, 0, QTableWidgetItem(h_key))
            self.table_haftalik.setItem(r, 1, QTableWidgetItem(f"{et:.3f} kg".replace('.', ',')))
            self.table_haftalik.setItem(r, 2, QTableWidgetItem(f"{tavuk:.3f} kg".replace('.', ',')))
            self.table_haftalik.setItem(r, 3, QTableWidgetItem(f"{(et+tavuk):.3f} kg".replace('.', ',')))
            zayi_item = QTableWidgetItem(f"{zayi:.3f} kg".replace('.', ','))
            if zayi > 0: zayi_item.setForeground(QColor("#F9E2AF"))
            self.table_haftalik.setItem(r, 4, zayi_item)
            self.table_haftalik.setItem(r, 5, QTableWidgetItem(_paket_str(self.haftalik_veri[h_key]["paket"])))

        for ay_key in sorted(self.aylik_veri.keys(), reverse=True):
            r = self.table_aylik.rowCount(); self.table_aylik.insertRow(r)
            et = self.aylik_veri[ay_key]["Et"]; tavuk = self.aylik_veri[ay_key]["Tavuk"]; zayi = self.aylik_veri[ay_key]["Zayi"]
            self.table_aylik.setItem(r, 0, QTableWidgetItem(ay_key))
            self.table_aylik.setItem(r, 1, QTableWidgetItem(f"{et:.3f} kg".replace('.', ',')))
            self.table_aylik.setItem(r, 2, QTableWidgetItem(f"{tavuk:.3f} kg".replace('.', ',')))
            self.table_aylik.setItem(r, 3, QTableWidgetItem(f"{(et+tavuk):.3f} kg".replace('.', ',')))
            zayi_item = QTableWidgetItem(f"{zayi:.3f} kg".replace('.', ','))
            if zayi > 0: zayi_item.setForeground(QColor("#F9E2AF"))
            self.table_aylik.setItem(r, 4, zayi_item)
            self.table_aylik.setItem(r, 5, QTableWidgetItem(_paket_str(self.aylik_veri[ay_key]["paket"])))

        def _fmt(adet, kg):
            return f"{adet} / {kg:.3f} kg".replace('.', ',') if adet > 0 else "-"
        for (gun_key, kat) in sorted(self.gelen_koli_veri.keys(), key=lambda x: datetime.strptime(x[0], "%d.%m.%Y"), reverse=True):
            d = self.gelen_koli_veri[(gun_key, kat)]
            r = self.table_gelen.rowCount(); self.table_gelen.insertRow(r)
            toplam_kg = d["4_kg"]+d["6_kg"]+d["8_kg"]+d["10_kg"]+d["12_kg"]+d["15_kg"]+d["20_kg"]+d["30_kg"]+d["40_kg"]+d["50_kg"]+d["diger_kg"]
            kat_item = QTableWidgetItem(kat)
            kat_item.setForeground(QColor("#F38BA8") if kat == "Et" else QColor("#F9E2AF"))
            self.table_gelen.setItem(r, 0, QTableWidgetItem(gun_key))
            self.table_gelen.setItem(r, 1, kat_item)
            self.table_gelen.setItem(r, 2,  QTableWidgetItem(_fmt(d["4_adet"],    d["4_kg"])))
            self.table_gelen.setItem(r, 3,  QTableWidgetItem(_fmt(d["6_adet"],    d["6_kg"])))
            self.table_gelen.setItem(r, 4,  QTableWidgetItem(_fmt(d["8_adet"],    d["8_kg"])))
            self.table_gelen.setItem(r, 5,  QTableWidgetItem(_fmt(d["10_adet"],   d["10_kg"])))
            self.table_gelen.setItem(r, 6,  QTableWidgetItem(_fmt(d["12_adet"],   d["12_kg"])))
            self.table_gelen.setItem(r, 7,  QTableWidgetItem(_fmt(d["15_adet"],   d["15_kg"])))
            self.table_gelen.setItem(r, 8,  QTableWidgetItem(_fmt(d["20_adet"],   d["20_kg"])))
            self.table_gelen.setItem(r, 9,  QTableWidgetItem(_fmt(d["30_adet"],   d["30_kg"])))
            self.table_gelen.setItem(r, 10, QTableWidgetItem(_fmt(d["40_adet"],   d["40_kg"])))
            self.table_gelen.setItem(r, 11, QTableWidgetItem(_fmt(d["50_adet"],   d["50_kg"])))
            self.table_gelen.setItem(r, 12, QTableWidgetItem(_fmt(d["diger_adet"], d["diger_kg"])))
            self.table_gelen.setItem(r, 13, QTableWidgetItem(f"{toplam_kg:.3f} kg".replace('.', ',')))

    def _stok_ozet_doldur(self, veriler):
        """Stok Özeti sekmesini doldur"""
        self.table_stok_ozet.setRowCount(0)
        ozet = {
            "Et":    {"giren": 0.0, "kalan": 0.0, "tuketilen": 0.0, "zayi": 0.0},
            "Tavuk": {"giren": 0.0, "kalan": 0.0, "tuketilen": 0.0, "zayi": 0.0},
        }
        for u in veriler:
            kat = u.get("kategori", "")
            if kat not in ozet: continue
            ilk = u.get("ilk_miktar", 0) or 0
            kal = u.get("kalan_miktar", 0) or 0
            zayi = u.get("zayi_miktar", 0) or 0
            trf = u.get("transfer_miktar", 0) or 0
            tuk = max(0, ilk - kal - zayi - trf)
            ozet[kat]["giren"] += ilk
            ozet[kat]["kalan"] += kal
            ozet[kat]["tuketilen"] += tuk
            ozet[kat]["zayi"] += zayi

        for kat, d in ozet.items():
            r = self.table_stok_ozet.rowCount()
            self.table_stok_ozet.insertRow(r)
            oran = (d["tuketilen"] / d["giren"] * 100) if d["giren"] > 0 else 0
            renk = QColor("#F38BA8") if kat == "Et" else QColor("#F9E2AF")
            items = [
                QTableWidgetItem(kat),
                QTableWidgetItem(f"{d['giren']:.3f} kg".replace('.', ',')),
                QTableWidgetItem(f"{d['kalan']:.3f} kg".replace('.', ',')),
                QTableWidgetItem(f"{d['tuketilen']:.3f} kg".replace('.', ',')),
                QTableWidgetItem(f"{d['zayi']:.3f} kg".replace('.', ',')),
                QTableWidgetItem(f"% {oran:.1f}".replace('.', ',')),
            ]
            for col, item in enumerate(items):
                item.setForeground(renk)
                self.table_stok_ozet.setItem(r, col, item)

        # Toplam satırı
        r = self.table_stok_ozet.rowCount()
        self.table_stok_ozet.insertRow(r)
        top_giren = sum(d["giren"] for d in ozet.values())
        top_kalan = sum(d["kalan"] for d in ozet.values())
        top_tuk = sum(d["tuketilen"] for d in ozet.values())
        top_zayi = sum(d["zayi"] for d in ozet.values())
        top_oran = (top_tuk / top_giren * 100) if top_giren > 0 else 0
        toplam_items = [
            QTableWidgetItem("TOPLAM"),
            QTableWidgetItem(f"{top_giren:.3f} kg".replace('.', ',')),
            QTableWidgetItem(f"{top_kalan:.3f} kg".replace('.', ',')),
            QTableWidgetItem(f"{top_tuk:.3f} kg".replace('.', ',')),
            QTableWidgetItem(f"{top_zayi:.3f} kg".replace('.', ',')),
            QTableWidgetItem(f"% {top_oran:.1f}".replace('.', ',')),
        ]
        for col, item in enumerate(toplam_items):
            item.setBackground(QColor("#313244"))
            item.setForeground(QColor("#CDD6F4"))
            f = item.font(); f.setBold(True); item.setFont(f)
            self.table_stok_ozet.setItem(r, col, item)

        # ── Paket Raporu tablosu ──────────────────────────────────────────
        PK_SIRASI = ["4","6","8","10","12","15","20","30","40","50","diger"]

        def _pk_dag_str(gun_pk_dict, kat):
            """'6kgx1, 15kgx2' formatında string döner"""
            parcalar = []
            for pk in PK_SIRASI:
                adet = gun_pk_dict.get(pk, {}).get(f"{kat}_adet", 0)
                if adet > 0:
                    etiket = "Diğer" if pk == "diger" else f"{pk}kg"
                    parcalar.append(f"{etiket}x{adet}")
            return ", ".join(parcalar) if parcalar else "-"

        gun_pk_agg = defaultdict(lambda: {pk: {"Et_adet":0,"Tavuk_adet":0} for pk in PK_SIRASI})
        for (gun_str, pk), d in self.paket_gun_veri.items():
            gun_pk_agg[gun_str][pk]["Et_adet"]    += d.get("Et_adet", 0)
            gun_pk_agg[gun_str][pk]["Tavuk_adet"] += d.get("Tavuk_adet", 0)

        self.table_paket_rapor.setRowCount(0)
        for gun_key in sorted(gun_pk_agg.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"), reverse=True):
            r = self.table_paket_rapor.rowCount(); self.table_paket_rapor.insertRow(r)
            self.table_paket_rapor.setItem(r, 0, QTableWidgetItem(gun_key))
            et_item = QTableWidgetItem(_pk_dag_str(gun_pk_agg[gun_key], "Et"))
            et_item.setForeground(QColor("#F38BA8"))
            self.table_paket_rapor.setItem(r, 1, et_item)
            tavuk_item = QTableWidgetItem(_pk_dag_str(gun_pk_agg[gun_key], "Tavuk"))
            tavuk_item.setForeground(QColor("#F9E2AF"))
            self.table_paket_rapor.setItem(r, 2, tavuk_item)

    def hizli_hucre_guncelle(self, satir, sutun, yeni_deger):
        tablo = self.aktif_tabloyu_al()
        if not tablo: return
        id_item = tablo.item(satir, 12)
        if not id_item: return
        urun_id = int(id_item.text())

        alan_adi = {0: "barkod", 1: "gelis_tarihi", 2: "kullanim_tarihi",
                    3: "kuvet_kullanim_tarihi", 4: "kuvet_miktar",
                    5: "takoz_kullanim_tarihi", 6: "takoz_miktar",
                    8: "kalan_miktar", 10: "zayi_miktar",
                    11: "zayi_tarihi"}.get(sutun)
        if not alan_adi: return

        if sutun in [4, 6, 8, 10]:
            try: yeni_deger = float(str(yeni_deger).replace("kg", "").replace(",", ".").strip())
            except ValueError: return
        else:
            if str(yeni_deger).strip() == "": yeni_deger = "-"

        guncelleme = {alan_adi: yeni_deger}

        # Küvet, takoz veya zayi girilince kalan otomatik hesapla
        if sutun in [4, 6, 10]:
            try:
                ilk   = float(str(tablo.item(satir, 7).text()).replace("kg", "").replace(",", ".").strip())
                kuvet = float(str(tablo.item(satir, 4).text()).replace("kg", "").replace(",", ".").strip()) if sutun != 4  else yeni_deger
                takoz = float(str(tablo.item(satir, 6).text()).replace("kg", "").replace(",", ".").strip()) if sutun != 6  else yeni_deger
                zayi  = float(str(tablo.item(satir, 10).text()).replace("kg", "").replace(",", ".").strip()) if sutun != 10 else yeni_deger
                kalan = max(0.0, ilk - kuvet - takoz - zayi)
                guncelleme["kalan_miktar"] = kalan
            except: pass

        basarili = self.api.urun_guncelle(urun_id, guncelleme)
        if not basarili:
            QMessageBox.warning(self, "Uyarı", "Güncelleme başarısız! Barkod zaten kayıtlı olabilir.")

    def hucre_degisti(self, item):
        tablo = item.tableWidget()
        sutun = item.column()
        secili_hucreler = tablo.selectedItems()
        ayni_sutundaki = [h for h in secili_hucreler if h.column() == sutun]
        tablo.blockSignals(True)
        if len(ayni_sutundaki) > 1 and item in ayni_sutundaki:
            for h in ayni_sutundaki:
                self.hizli_hucre_guncelle(h.row(), sutun, item.text())
        else:
            self.hizli_hucre_guncelle(item.row(), sutun, item.text())
        tablo.blockSignals(False)
        self.verileri_yukle()

    def yeni_urun_penceresi_ac(self):
        # Aktif sekmeye göre kategori belirle
        idx = self.tabs.currentIndex()
        if idx == 1:
            kat = "Et"
        elif idx == 2:
            kat = "Tavuk"
        else:
            kat = None
        dialog = UrunEkleDialog(self, varsayilan_kategori=kat)
        dialog.setStyleSheet(self.styleSheet())
        if dialog.exec() == QDialog.DialogCode.Accepted:
            barkod, kategori, gelis, kullanim, tekrar, ilk_mik, kalan_mik = dialog.verileri_al()
            if barkod.strip():
                try:
                    self.api.urun_ekle({
                        "barkod": barkod, "kategori": kategori,
                        "gelis_tarihi": gelis, "kullanim_tarihi": kullanim,
                        "tekrar_kullanim_tarihi": tekrar,
                        "ilk_miktar": ilk_mik, "kalan_miktar": kalan_mik
                    })
                    self.verileri_yukle()
                except Exception as e:
                    if "kayıtlı" in str(e) or "409" in str(e):
                        QMessageBox.warning(self, "Hata", "Bu barkod numarası zaten sistemde kayıtlı!")
                    else:
                        QMessageBox.critical(self, "Hata", str(e))

    def urun_sil(self):
        tablo = self.aktif_tabloyu_al()
        if not tablo: return
        secili_ogeler = tablo.selectedItems()
        if not secili_ogeler: return
        satir_indeksleri = set(item.row() for item in secili_ogeler)
        cevap = QMessageBox.question(self, "Onay", f"Seçili {len(satir_indeksleri)} kaydı silmek istediğinize emin misiniz?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if cevap == QMessageBox.StandardButton.Yes:
            for i in satir_indeksleri:
                self.api.urun_sil(int(tablo.item(i, 12).text()))
            self.verileri_yukle()

    def tarih_formatla(self, val):
        if isinstance(val, datetime): return val.strftime("%d.%m.%Y")
        elif val:
            metin = str(val).strip()
            if "-" in metin and len(metin) >= 10:
                try: return datetime.strptime(metin[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
                except: pass
            return metin
        return "-"

    def excel_disa_aktar(self):
        dosya_yolu, _ = QFileDialog.getSaveFileName(self, "Excel'e Aktar", f"Stok_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel (*.xlsx)")
        if dosya_yolu:
            try:
                veriler = self.api.urunleri_getir()
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Stok Verileri"
                ws.append(["ID", "Barkod", "Kategori", "Geliş T.", "Kullanım T.", "Tekrar Kullanım", "İlk Miktar", "Harcanan", "Kalan Miktar"])
                for u in veriler:
                    ws.append([u["id"], u["barkod"], u["kategori"], u["gelis_tarihi"],
                                u.get("kullanim_tarihi", "-"), u.get("tekrar_kullanim_tarihi", "-"),
                                u["ilk_miktar"], u.get("tekrar_miktar", 0), u["kalan_miktar"]])
                wb.save(dosya_yolu)
                QMessageBox.information(self, "Başarılı", "Excel dosyası oluşturuldu!")
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))

    def analiz_raporu_disa_aktar(self):
        dosya_yolu, _ = QFileDialog.getSaveFileName(self, "Raporu Kaydet", f"Analiz_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel (*.xlsx)")
        if not dosya_yolu: return
        try:
            wb = openpyxl.Workbook()
            ws_g = wb.active; ws_g.title = "Günlük Çıkışlar"
            ws_g.append(["Günlük Periyot", "Et (kg)", "Tavuk (kg)", "Toplam (kg)"])
            for k in sorted(self.gunluk_veri.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"), reverse=True):
                et = self.gunluk_veri[k]["Et"]; tavuk = self.gunluk_veri[k]["Tavuk"]
                ws_g.append([k, et, tavuk, et + tavuk])
            ws_h = wb.create_sheet("Haftalık Çıkışlar")
            ws_h.append(["Haftalık Periyot", "Et (kg)", "Tavuk (kg)", "Toplam (kg)"])
            for k in sorted(self.haftalik_veri.keys(), reverse=True):
                et = self.haftalik_veri[k]["Et"]; tavuk = self.haftalik_veri[k]["Tavuk"]
                ws_h.append([k, et, tavuk, et + tavuk])
            ws_a = wb.create_sheet("Aylık Çıkışlar")
            ws_a.append(["Aylık Periyot", "Et (kg)", "Tavuk (kg)", "Toplam (kg)"])
            for k in sorted(self.aylik_veri.keys(), reverse=True):
                et = self.aylik_veri[k]["Et"]; tavuk = self.aylik_veri[k]["Tavuk"]
                ws_a.append([k, et, tavuk, et + tavuk])
            wb.save(dosya_yolu)
            QMessageBox.information(self, "Başarılı", "Analiz raporu oluşturuldu!")
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))

    def excelden_iceri_aktar(self):
        dosya_yolu, _ = QFileDialog.getOpenFileName(self, "Excel Seç", "", "Excel (*.xlsx)")
        if dosya_yolu:
            try:
                wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
                basarili = hatali = 0
                for sayfa in wb.sheetnames:
                    sheet = wb[sayfa]
                    kat = "Tavuk" if "TAVUK" in sayfa.upper() else "Et"
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not row[0]: continue
                        barkod = str(row[0]).strip()
                        if "TOPLAM" in barkod.upper(): continue
                        gelis = self.tarih_formatla(row[1])
                        try: ilk_mik = float(str(row[2]).replace(',', '.')) if row[2] else 0.0
                        except: ilk_mik = 0.0
                        kullanim = self.tarih_formatla(row[3])
                        try: kalan_mik = float(str(row[4]).replace(',', '.')) if row[4] else 0.0
                        except: kalan_mik = 0.0
                        tekrar = self.tarih_formatla(row[6]) if len(row) > 6 else "-"
                        try:
                            self.api.urun_ekle({"barkod": barkod, "kategori": kat, "gelis_tarihi": gelis,
                                                "kullanim_tarihi": kullanim, "tekrar_kullanim_tarihi": tekrar,
                                                "ilk_miktar": ilk_mik, "kalan_miktar": kalan_mik})
                            basarili += 1
                        except: hatali += 1
                self.verileri_yukle()
                QMessageBox.information(self, "Sonuç", f"Eklendi: {basarili}\nAtlanan: {hatali}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))


# ─────────────────────────────────────────
#  GİRİŞ EKRANI
# ─────────────────────────────────────────
class GirisEkrani(QDialog):
    AYAR_DOSYASI = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "giris_ayar.ini")

    def __init__(self):
        super().__init__()
        self.api = ApiIstemci()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.resize(350, 460)
        self.setWindowIcon(uygulama_ikonu_yukle())

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.ana_widget = QWidget()
        self.ana_widget.setStyleSheet("QWidget { background-color: #181825; border: 2px solid #313244; border-radius: 15px; }")
        self.ana_layout = QVBoxLayout(self.ana_widget)
        self.ana_layout.setContentsMargins(20, 20, 20, 20)

        ust = QHBoxLayout()
        ust.addStretch()
        kapat = QPushButton("✕")
        kapat.setStyleSheet("QPushButton { background-color: transparent; color: #F38BA8; font-weight: bold; font-size: 16px; border: none; } QPushButton:hover { color: #ff0000; }")
        kapat.clicked.connect(self.reject)
        ust.addWidget(kapat)
        self.ana_layout.addLayout(ust)

        ikon = QLabel("☁️🔒")
        ikon.setStyleSheet("font-size: 55px; border: none; background: transparent;")
        ikon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ana_layout.addWidget(ikon)

        baslik = QLabel("StockFlow")
        baslik.setStyleSheet("color: #CDD6F4; font-size: 20px; font-weight: bold; border: none; background: transparent;")
        baslik.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ana_layout.addWidget(baslik)
        self.ana_layout.addSpacing(15)

        self.sifre_input = QLineEdit()
        self.sifre_input.setPlaceholderText("Şifrenizi girin...")
        self.sifre_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.sifre_input.setStyleSheet("background-color: #1E1E2E; color: #CDD6F4; border: 2px solid #313244; border-radius: 10px; padding: 12px; font-size: 14px;")
        self.ana_layout.addWidget(self.sifre_input)
        self.ana_layout.addSpacing(8)

        # Beni hatırla checkbox
        self.hatirla_cb = QCheckBox("Beni hatırla")
        self.hatirla_cb.setStyleSheet("""
            QCheckBox { color: #A6ADC8; font-size: 12px; border: none; background: transparent; }
            QCheckBox::indicator { width: 16px; height: 16px; border-radius: 4px; border: 2px solid #45475A; background: #1E1E2E; }
            QCheckBox::indicator:checked { background: #FAB387; border: 2px solid #FAB387; }
        """)
        self.ana_layout.addWidget(self.hatirla_cb)
        self.ana_layout.addSpacing(8)

        self.giris_btn = QPushButton("Giriş Yap")
        self.giris_btn.setStyleSheet("background-color: #FAB387; color: #11111B; border-radius: 10px; padding: 12px; font-weight: bold; font-size: 15px;")
        self.giris_btn.clicked.connect(self.giris_kontrol)
        self.ana_layout.addWidget(self.giris_btn)

        self.durum_lbl = QLabel("")
        self.durum_lbl.setStyleSheet("color: #F38BA8; font-size: 12px; border: none; background: transparent;")
        self.durum_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.ana_layout.addWidget(self.durum_lbl)

        self.ana_layout.addStretch()
        self.layout.addWidget(self.ana_widget)
        self.sifre_input.returnPressed.connect(self.giris_kontrol)

        # Kayıtlı şifreyi yükle
        self._kayitli_yukle()

    def _kayitli_yukle(self):
        try:
            ayar = QSettings(self.AYAR_DOSYASI, QSettings.Format.IniFormat)
            if ayar.value("hatirla", False, type=bool):
                import base64
                sifre = base64.b64decode(ayar.value("sifre", "")).decode()
                self.sifre_input.setText(sifre)
                self.hatirla_cb.setChecked(True)
        except: pass

    def _kayitli_kaydet(self, sifre):
        try:
            import base64
            ayar = QSettings(self.AYAR_DOSYASI, QSettings.Format.IniFormat)
            if self.hatirla_cb.isChecked():
                ayar.setValue("hatirla", True)
                ayar.setValue("sifre", base64.b64encode(sifre.encode()).decode())
            else:
                ayar.setValue("hatirla", False)
                ayar.remove("sifre")
            ayar.sync()
        except: pass

    def otomatik_giris(self):
        """Kayıtlı şifre varsa direkt giriş yap"""
        try:
            ayar = QSettings(self.AYAR_DOSYASI, QSettings.Format.IniFormat)
            if ayar.value("hatirla", False, type=bool):
                import base64
                sifre = base64.b64decode(ayar.value("sifre", "")).decode()
                if sifre:
                    self.giris_btn.setText("Bağlanıyor...")
                    self.giris_btn.setEnabled(False)
                    self.durum_lbl.setText("Otomatik giriş yapılıyor...")
                    QApplication.processEvents()
                    basarili = self.api.giris_yap(sifre)
                    if basarili:
                        self.accept()
                        return True
                    else:
                        self.giris_btn.setText("Giriş Yap")
                        self.giris_btn.setEnabled(True)
                        self.durum_lbl.setText("")
        except Exception:
            pass
        return False

    def giris_kontrol(self):
        self.giris_btn.setText("Bağlanıyor...")
        self.giris_btn.setEnabled(False)
        QApplication.processEvents()

        sifre = self.sifre_input.text()
        basarili = self.api.giris_yap(sifre)

        if basarili:
            self._kayitli_kaydet(sifre)
            self.accept()
        else:
            self.durum_lbl.setText("Hatalı şifre veya sunucuya ulaşılamadı!")
            self.sifre_input.clear()
            self.giris_btn.setText("Giriş Yap")
            self.giris_btn.setEnabled(True)


# ─────────────────────────────────────────
#  UYGULAMA BAŞLANGICI
# ─────────────────────────────────────────
if __name__ == "__main__":
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('stockflow.stok_sistemi.v15')
    except Exception:
        pass

    # DPI ölçeklendirme — farklı ekran çözünürlüklerine adaptif
    try:
        from PyQt6.QtCore import Qt as _Qt
        QApplication.setHighDpiScaleFactorRoundingPolicy(
            _Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
        )
    except Exception:
        pass

    app = QApplication(sys.argv)
    app.setWindowIcon(uygulama_ikonu_yukle())

    # Ekran DPI'ına göre global font boyutunu ayarla
    ekran = app.primaryScreen()
    dpi = ekran.logicalDotsPerInch()
    if dpi <= 96:       baz_font = 11   # 1080p normal
    elif dpi <= 120:    baz_font = 12   # 1080p %125
    elif dpi <= 144:    baz_font = 13   # 4K %150
    else:               baz_font = 14   # 4K %175+
    app.setFont(QFont("Arial", baz_font))

    # ── SPLASH EKRANI ──
    from PyQt6.QtGui import QPixmap, QPainter, QLinearGradient, QColor, QFont
    from PyQt6.QtCore import Qt, QRect

    splash_px = QPixmap(480, 280)
    splash_px.fill(QColor("#181825"))

    painter = QPainter(splash_px)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)

    # Arka plan gradient
    grad = QLinearGradient(0, 0, 480, 280)
    grad.setColorAt(0.0, QColor("#1E1E2E"))
    grad.setColorAt(1.0, QColor("#11111B"))
    painter.fillRect(QRect(0, 0, 480, 280), grad)

    # Üst turuncu şerit
    painter.fillRect(QRect(0, 0, 480, 4), QColor("#FAB387"))

    # İkon
    painter.setFont(QFont("Segoe UI Emoji", 52))
    painter.setPen(QColor("#FAB387"))
    painter.drawText(QRect(0, 30, 480, 90), Qt.AlignmentFlag.AlignHCenter, "🐱")

    # Başlık
    painter.setFont(QFont("Arial", 22, QFont.Weight.Bold))
    painter.setPen(QColor("#CDD6F4"))
    painter.drawText(QRect(0, 120, 480, 45), Qt.AlignmentFlag.AlignHCenter, "StockFlow")

    # Alt yazı
    painter.setFont(QFont("Arial", 11))
    painter.setPen(QColor("#A6ADC8"))
    painter.drawText(QRect(0, 165, 480, 30), Qt.AlignmentFlag.AlignHCenter, "Et & Tavuk Stok Yönetimi")

    # Yükleniyor yazısı
    painter.setFont(QFont("Arial", 10))
    painter.setPen(QColor("#6C7086"))
    painter.drawText(QRect(0, 240, 480, 25), Qt.AlignmentFlag.AlignHCenter, "Başlatılıyor...")

    # Alt turuncu şerit
    painter.fillRect(QRect(0, 276, 480, 4), QColor("#FAB387"))

    painter.end()

    splash = QSplashScreen(splash_px, Qt.WindowType.WindowStaysOnTopHint)
    splash.setWindowFlag(Qt.WindowType.FramelessWindowHint)
    splash.show()
    app.processEvents()

    # 2 saniye göster
    import time
    time.sleep(2)
    splash.close()

    giris = GirisEkrani()
    # Kayıtlı şifre varsa otomatik giriş
    if not giris.otomatik_giris():
        if giris.exec() != QDialog.DialogCode.Accepted:
            sys.exit()
    if True:
        window = StokSistemi(giris.api)
        window.show()
        sys.exit(app.exec())
    else:
        sys.exit()
